from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


TRACKING_SHEET = "Tracking Sheet"
TEST_TRACK_SHEET = "Test Track Table"
SYSTEM_SHEET = "System Table"
CARTRIDGE_SHEET = "Cartridge Table"


class WorkbookImportError(RuntimeError):
    pass


def export_tracking_workbook_to_json_fixtures(
    workbook_path: str | Path,
    output_dir: str | Path,
    file_stems: Iterable[str] | None = None,
) -> dict[str, Path]:
    workbook = load_workbook(workbook_path, data_only=True)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    file_stem_filter = None if file_stems is None else {stem.lower() for stem in file_stems}

    acquisitions = _parse_acquisitions(workbook[TRACKING_SHEET], file_stem_filter)
    needed_test_tracks = sorted({record["test_track_name"] for record in acquisitions if record.get("test_track_name")})
    needed_systems = sorted({record["system_id"] for record in acquisitions if record.get("system_id") is not None})
    needed_cartridges = sorted({record["cartridge_name"] for record in acquisitions if record.get("cartridge_name")})

    test_tracks = _parse_test_tracks(workbook[TEST_TRACK_SHEET], set(needed_test_tracks))
    systems = _parse_systems(workbook[SYSTEM_SHEET], set(needed_systems))
    cartridges = _parse_cartridges(workbook[CARTRIDGE_SHEET], set(needed_cartridges))

    outputs = {
        "acquisitions_json": output_path / "acquisitions.json",
        "test_tracks_json": output_path / "test_tracks.json",
        "systems_json": output_path / "systems.json",
        "cartridges_json": output_path / "cartridges.json",
    }

    outputs["acquisitions_json"].write_text(json.dumps(acquisitions, indent=2, default=str), encoding="utf-8")
    outputs["test_tracks_json"].write_text(json.dumps(test_tracks, indent=2, default=str), encoding="utf-8")
    outputs["systems_json"].write_text(json.dumps(systems, indent=2, default=str), encoding="utf-8")
    outputs["cartridges_json"].write_text(json.dumps(cartridges, indent=2, default=str), encoding="utf-8")
    return outputs


def _parse_acquisitions(ws, file_stem_filter: set[str] | None) -> list[dict]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {str(value): idx for idx, value in enumerate(header_row) if value is not None}

    required = [
        "File (From J.R.)",
        "Acq Date",
        "System",
        "Cartridge",
        "Wally Zenith",
        "Stylus ZE",
        "Eff Len",
        "Offset Ang",
        "Over hang",
        "Request Ovrhang",
        "∆ Ovrhang - Actual",
        "∆ P to S - Actual",
        "P to S - Actual",
        "Test Track",
        "Digit-izer",
        "Comments",
    ]
    for column in required:
        if column not in header_index:
            raise WorkbookImportError(f"Missing acquisition column: {column}")

    records: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        file_stem = row[header_index["File (From J.R.)"]]
        if not file_stem:
            continue
        file_stem = str(file_stem)
        if file_stem_filter is not None and file_stem.lower() not in file_stem_filter:
            continue
        records.append(
            {
                "file_stem": file_stem,
                "recorded_at": _string_or_none(row[header_index["Acq Date"]]),
                "digitizer": _string_or_none(row[header_index["Digit-izer"]]),
                "test_track_name": _string_or_none(row[header_index["Test Track"]]),
                "system_id": _float_or_none(row[header_index["System"]]),
                "cartridge_name": _string_or_none(row[header_index["Cartridge"]]),
                "cantilever_yaw_deg": _float_or_none(row[header_index["Wally Zenith"]]),
                "stylus_yaw_deg": _float_or_none(row[header_index["Stylus ZE"]]),
                "effective_length_mm": _float_or_none(row[header_index["Eff Len"]]),
                "offset_angle_deg": _float_or_none(row[header_index["Offset Ang"]]),
                "overhang_mm": _float_or_none(row[header_index["Over hang"]]),
                "required_overhang_mm": _float_or_none(row[header_index["Request Ovrhang"]]),
                "overhang_adjustment_mm": _float_or_none(row[header_index["∆ Ovrhang - Actual"]]),
                "pivot_spindle_adjustment_mm": _float_or_none(row[header_index["∆ P to S - Actual"]]),
                "actual_pivot_to_spindle_mm": _float_or_none(row[header_index["P to S - Actual"]]),
                "comments": _string_or_none(row[header_index["Comments"]]),
            }
        )
    return records


def _parse_test_tracks(ws, needed_names: set[str]) -> list[dict]:
    header_row_idx = 3
    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    header_index = {str(value): idx for idx, value in enumerate(header_row) if value is not None}
    required = ["Name", "Outer radius", "Inner radius", "Notes"]
    for column in required:
        if column not in header_index:
            raise WorkbookImportError(f"Missing test track column: {column}")

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[header_index["Name"]]
        if not name:
            continue
        name = str(name)
        if needed_names and name not in needed_names:
            continue
        records.append(
            {
                "name": name,
                "outer_radius_mm": float(row[header_index["Outer radius"]]),
                "inner_radius_mm": float(row[header_index["Inner radius"]]),
                "notes": _string_or_none(row[header_index["Notes"]]),
            }
        )
    return records


def _parse_systems(ws, needed_ids: set[float]) -> list[dict]:
    header_row_idx = 3
    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    header_index = {str(value): idx for idx, value in enumerate(header_row) if value is not None}
    required = ["System Name", "Turntable", "Tonearm", "Headshell", "Shim Type", "Isolation", "Notes"]
    for column in required:
        if column not in header_index:
            raise WorkbookImportError(f"Missing system column: {column}")

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        system_id = row[header_index["System Name"]]
        if system_id is None:
            continue
        system_id = float(system_id)
        if needed_ids and system_id not in needed_ids:
            continue
        records.append(
            {
                "system_id": system_id,
                "turntable": _string_or_none(row[header_index["Turntable"]]),
                "tonearm": _string_or_none(row[header_index["Tonearm"]]),
                "headshell": _string_or_none(row[header_index["Headshell"]]),
                "shim": _string_or_none(row[header_index["Shim Type"]]),
                "isolation": _string_or_none(row[header_index["Isolation"]]),
                "notes": _string_or_none(row[header_index["Notes"]]),
            }
        )
    return records


def _parse_cartridges(ws, needed_names: set[str]) -> list[dict]:
    header_row_idx = 3
    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    header_index = {str(value): idx for idx, value in enumerate(header_row) if value is not None}
    required = ["Cartridge", "L-R Contact Dim", "Zenith Error*", "WZ Alignment", "SRA - corrected", "VTA - corrected", "Azimuth**", "Notes"]
    for column in required:
        if column not in header_index:
            raise WorkbookImportError(f"Missing cartridge column: {column}")

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[header_index["Cartridge"]]
        if not name:
            continue
        name = str(name)
        if needed_names and name not in needed_names:
            continue
        records.append(
            {
                "cartridge_name": name,
                "lr_um": _float_or_none(row[header_index["L-R Contact Dim"]]),
                "ze_deg": _float_or_none(row[header_index["Zenith Error*"]]),
                "wally_zenith_deg": _float_or_none(row[header_index["WZ Alignment"]]),
                "sra_deg": _float_or_none(row[header_index["SRA - corrected"]]),
                "vta_deg": _float_or_none(row[header_index["VTA - corrected"]]),
                "notes": _string_or_none(row[header_index["Notes"]]),
            }
        )
    return records


def _float_or_none(value):
    if value is None or value == "":
        return None
    return float(value)


def _string_or_none(value):
    if value is None or value == "":
        return None
    return str(value)
