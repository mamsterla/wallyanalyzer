from __future__ import annotations

import csv
import math
import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable

import requests

BIGLOBE_ARMDATA_URL = "https://www7a.biglobe.ne.jp/~yosh/armdata.htm"
BIGLOBE_EXCLUDED_MODELS: dict[tuple[str, str], str] = {
    ("jml-co", "ta-3a"): "Excluded from BIGLOBE import: period source trail points to JML Company as an alignment-protractor vendor, not a confirmed tonearm manufacturer.",
}
WORKBOOK_MANUFACTURER_ALIASES: dict[str, str] = {
    "basis": "Basis Audio",
    "breuer-dynamic": "Breuer",
    "korf-audio": "Korf",
    "luxman-saec": "Saec",
    "mitchell": "Michell",
    "morch": "Moerch",
    "m-rch": "Moerch",
    "nottingham": "Nottingham Analogue",
    "project": "Pro-Ject",
    "the-wand": "Wand",
    "triplanar": "Tri-Planar",
    "vertere": "Vertere Acoustics",
}
WORKBOOK_EXCLUDED_MODELS: dict[tuple[str, str], str] = {
    ("avid", "acutus"): "Workbook row describes turntable, not standalone tonearm.",
    ("dual", "721"): "Workbook row describes turntable, not standalone tonearm.",
    ("music-hall", "2-5"): "Workbook row describes turntable, not standalone tonearm.",
    ("music-hall", "3-3"): "Workbook row describes turntable, not standalone tonearm.",
    ("music-hall", "5-3"): "Workbook row describes turntable, not standalone tonearm.",
    ("music-hall", "7-3"): "Workbook row describes turntable, not standalone tonearm.",
    ("pro-ject", "carbon-debut"): "Workbook row describes turntable, not standalone tonearm.",
    ("pro-ject", "rpm-carbon-10"): "Workbook row describes turntable, not standalone tonearm.",
    ("pro-ject", "signature-10"): "Workbook row describes turntable, not standalone tonearm.",
    ("pro-ject", "xtension-10-evolution"): "Workbook row describes turntable, not standalone tonearm.",
    ("pro-ject", "xtension-9-evolution"): "Workbook row describes turntable, not standalone tonearm.",
    ("pro-ject", "xtension-12-evolution"): "Workbook row describes turntable, not standalone tonearm.",
    ("technics", "sl-1000r"): "Workbook row describes turntable, not standalone tonearm.",
    ("technics", "sl-1200g"): "Workbook row describes turntable, not standalone tonearm.",
    ("technics", "sl-1200gr"): "Workbook row describes turntable, not standalone tonearm.",
    ("technics", "sl-1210gr"): "Workbook row describes turntable, not standalone tonearm.",
    ("technics", "sl-1500c"): "Workbook row describes turntable, not standalone tonearm.",
    ("vertere-acoustics", "sg-1"): "Workbook row describes turntable, not standalone tonearm.",
    ("vpi", "player"): "Workbook row describes turntable, not standalone tonearm.",
}
NULL_ALIGNMENT_TYPE_VALUES = ("Stevenson", "Baewald", "Loefgren", "Unknown", "Other")
NULL_ALIGNMENT_REFERENCE_POINTS = {
    "Stevenson": (60.325, 117.42),
    "Baewald": (66.0, 120.9),
    "Loefgren": (70.285, 116.605),
}
NULL_ALIGNMENT_DISTANCE_TOLERANCE_MM = 1.5


MANUFACTURER_RESEARCH_SEEDS: list[dict[str, str]] = [
    {
        "manufacturer_name": "Tri-Planar",
        "canonical_match_name": "Tri-Planar",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "primary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Tri-Planar tonearm Mk VII U2 SE U12 manual effective length overhang",
        "notes": "Primary upgrade-arm coverage target.",
    },
    {
        "manufacturer_name": "Graham",
        "canonical_match_name": "Graham",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "primary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Graham tonearm Phantom manual effective length pivot to spindle",
        "notes": "Primary upgrade-arm coverage target.",
    },
    {
        "manufacturer_name": "Kuzma",
        "canonical_match_name": "Kuzma",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "primary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Kuzma tonearm 4Point Stogi Air Line manual effective length overhang",
        "notes": "Primary upgrade-arm coverage target.",
    },
    {
        "manufacturer_name": "Supatrac",
        "canonical_match_name": "Supatrac",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "primary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Supatrac Blackbird Nighthawk Farpoint manual effective length overhang",
        "notes": "Primary upgrade-arm coverage target.",
    },
    {
        "manufacturer_name": "Origin Live",
        "canonical_match_name": "Origin Live",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "primary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Origin Live tonearm Alliance Onyx Silver Zephyr manual overhang",
        "notes": "Primary upgrade-arm coverage target.",
    },
    {
        "manufacturer_name": "Rega",
        "canonical_match_name": "Rega",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "primary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Rega RB250 RB300 RB330 effective mass mounting distance overhang",
        "notes": "Primary upgrade-arm coverage target.",
    },
    {
        "manufacturer_name": "VPI",
        "canonical_match_name": "VPI",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "VPI JMW tonearm manual effective length overhang",
        "notes": "Secondary combo coverage target.",
    },
    {
        "manufacturer_name": "Pro-Ject",
        "canonical_match_name": "Pro-Ject",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Pro-Ject tonearm 9cc 10cc 12cc manual effective length overhang",
        "notes": "Secondary combo coverage target.",
    },
    {
        "manufacturer_name": "Music Hall",
        "canonical_match_name": "Music Hall",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Music Hall turntable tonearm specs manual effective length overhang",
        "notes": "Secondary combo coverage target.",
    },
    {
        "manufacturer_name": "TW-Acustic",
        "canonical_match_name": "TW-Acustic",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "TW-Acustic Raven tonearm manual effective length overhang",
        "notes": "Secondary combo coverage target.",
    },
    {
        "manufacturer_name": "Basis Audio",
        "canonical_match_name": "Basis Audio",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Basis Audio Vector tonearm manual effective length overhang",
        "notes": "Secondary combo coverage target.",
    },
    {
        "manufacturer_name": "Gold Note",
        "canonical_match_name": "Gold Note",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "Gold Note B-5.1 B-7 Ceramic tonearm geometry official",
        "notes": "Italian manufacturer added by user.",
    },
    {
        "manufacturer_name": "Ortofon",
        "canonical_match_name": "Ortofon",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "Ortofon AS-212R AS-309R RS-212D RS-309D tonearm geometry official",
        "notes": "Historic and current Ortofon tonearm line added by user.",
    },
    {
        "manufacturer_name": "Dr. Feickert",
        "canonical_match_name": "Dr. Feickert",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority",
        "search_terms": "Feickert Straight-10 tonearm effective length overhang mounting distance",
        "notes": "Secondary combo coverage target.",
    },
    {
        "manufacturer_name": "Audiomods",
        "canonical_match_name": "Audiomods",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-example",
        "search_terms": "Audiomods tonearm Series Six manual effective length overhang",
        "notes": "Cottage brand example; may take more manual research.",
    },
    {
        "manufacturer_name": "AMG",
        "canonical_match_name": "AMG",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "AMG 9W2 12JT tonearm geometry effective length overhang",
        "notes": "High-end German tonearm maker added by user.",
    },
    {
        "manufacturer_name": "Korf",
        "canonical_match_name": "Korf",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "Korf tonearm TA-SF9 TA-SF11R geometry effective length overhang",
        "notes": "High-end geometry-focused tonearm maker added by user.",
    },
    {
        "manufacturer_name": "Schroeder",
        "canonical_match_name": "Schroeder",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "Schroeder tonearm Reference CB No.2 DPS geometry mounting distance",
        "notes": "High-end boutique tonearm maker added by user.",
    },
    {
        "manufacturer_name": "Thales",
        "canonical_match_name": "Thales",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "Thales Statement Simplicity tonearm manual effective length",
        "notes": "Boutique reference-arm coverage target.",
    },
    {
        "manufacturer_name": "SAT",
        "canonical_match_name": "SAT",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "SAT LM-09 LM-12 tonearm geometry effective length overhang",
        "notes": "Swedish Analog Technologies reference-arm coverage target.",
    },
    {
        "manufacturer_name": "Well Tempered Lab",
        "canonical_match_name": "Well Tempered Lab",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "Well Tempered Lab LTD tonearm manual effective length",
        "notes": "Boutique reference-arm coverage target.",
    },
    {
        "manufacturer_name": "TechDAS",
        "canonical_match_name": "TechDAS",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-priority-followup",
        "search_terms": "TechDAS Air Force 10 tonearm effective length overhang",
        "notes": "High-end combo coverage target.",
    },
    {
        "manufacturer_name": "Acoustic Signature",
        "canonical_match_name": "Acoustic Signature",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Acoustic Signature TA-5000 NEO TA-7000 NEO tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Acoustical Systems",
        "canonical_match_name": "Acoustical Systems",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Acoustical Systems AQUILAR 10 AXIOM 12 reference tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Clearaudio",
        "canonical_match_name": "Clearaudio",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Clearaudio Tracer Black Carbon Fiber tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "EMT",
        "canonical_match_name": "EMT",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "EMT 909-HI 912 tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "J.Sikora",
        "canonical_match_name": "J.Sikora",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "J.Sikora KV12 VTA tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Linn",
        "canonical_match_name": "Linn",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Linn Ekos SE tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Schick",
        "canonical_match_name": "Schick",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Thomas Schick 10.5 12 tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Sorane",
        "canonical_match_name": "Sorane",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Sorane SA-1.2 tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "ViV Laboratory",
        "canonical_match_name": "ViV Laboratory",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "ViV Laboratory Rigid Float HA9 tonearm geometry underhung",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Wand",
        "canonical_match_name": "Wand",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "stereophile-2025-followup",
        "search_terms": "Wand Master 12 tonearm geometry effective length overhang",
        "notes": "Stereophile 2025 recommended tonearm maker follow-up.",
    },
    {
        "manufacturer_name": "Vertere Acoustics",
        "canonical_match_name": "Vertere Acoustics",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Vertere Acoustics tonearm geometry SG-1 PTA Groove Runner overhang effective length",
        "notes": "User-requested manufacturer follow-up.",
    },
    {
        "manufacturer_name": "Michell",
        "canonical_match_name": "Michell",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Michell tonearm geometry TecnoArm A TecnoArm 2 effective length overhang",
        "notes": "User-requested manufacturer follow-up.",
    },
    {
        "manufacturer_name": "Dynavector",
        "canonical_match_name": "Dynavector",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Dynavector tonearm DV-501 DV-505 DV-507 geometry effective length overhang",
        "notes": "User-requested manufacturer follow-up for legacy and exact arm coverage.",
    },
    {
        "manufacturer_name": "Audio-Technica",
        "canonical_match_name": "Audio-Technica",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Audio-Technica tonearm AT-1001 AT-1005II AT-1007 AT-1009 AT-1100 AT-1120 AT-1501 AT-1503 geometry",
        "notes": "User-requested manufacturer follow-up for historic Audio-Technica arm families.",
    },
    {
        "manufacturer_name": "J.Sikora",
        "canonical_match_name": "J.Sikora",
        "manufacturer_type": "turntable-combo",
        "coverage_focus": "high-end-turntable-combos",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "J.Sikora J. Sikora tonearm KV12 VTA geometry effective length overhang",
        "notes": "User-requested manufacturer follow-up; normalize punctuation variants to J.Sikora.",
    },
    {
        "manufacturer_name": "Fidelity Research",
        "canonical_match_name": "Fidelity Research",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Fidelity Research tonearm FR-12 FR-14 FR-24 FR-34 FR-54 FR-64 FR-66S FR-66fx geometry",
        "notes": "User-requested manufacturer follow-up for classic FR arm families.",
    },
    {
        "manufacturer_name": "Glanz",
        "canonical_match_name": "Glanz",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Glanz tonearm geometry MH-94S MH-104S MH-124S effective length overhang",
        "notes": "User-requested manufacturer follow-up.",
    },
    {
        "manufacturer_name": "Pear Audio",
        "canonical_match_name": "Pear Audio",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Pear Audio tonearm geometry Cornet tonearm effective length overhang",
        "notes": "User-requested manufacturer follow-up.",
    },
    {
        "manufacturer_name": "Audio Creative",
        "canonical_match_name": "Audio Creative",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "editorial-followup",
        "search_terms": "Audio Creative Groovemaster 3 4 tonearm geometry effective length mounting distance overhang",
        "notes": "Follow-up seeded from The Ear Groovemaster 4 review and current boutique-arm coverage expansion.",
    },
    {
        "manufacturer_name": "Acoustic Solid",
        "canonical_match_name": "Acoustic Solid",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Acoustic Solid tonearm WTB 213 WTB 370 geometry effective length overhang",
        "notes": "Missing from current DB/queue but present in broad external manufacturer list.",
    },
    {
        "manufacturer_name": "Air Tangent",
        "canonical_match_name": "Air Tangent",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "parallel-tracking-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Air Tangent tonearm 2A 2B geometry effective length linear tracker",
        "notes": "High-priority missing linear-tracking manufacturer from broad external list.",
    },
    {
        "manufacturer_name": "Rabco",
        "canonical_match_name": "Rabco",
        "manufacturer_type": "parallel-tracker",
        "coverage_focus": "parallel-tracking-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "user-manufacturer-followup",
        "search_terms": "Rabco SL-8 SL-8E tonearm manual geometry linear tracker",
        "notes": "User-requested linear-tracking manufacturer follow-up.",
    },
    {
        "manufacturer_name": "Alphason Designs",
        "canonical_match_name": "Alphason",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Alphason Designs HR-100S HR-100MCS tonearm geometry effective length overhang",
        "notes": "High-priority missing classic upgrade-arm maker from broad external list.",
    },
    {
        "manufacturer_name": "Audio Note",
        "canonical_match_name": "Audio Note",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Audio Note Arm One II Arm Three II tonearm geometry effective length overhang",
        "notes": "High-priority missing boutique manufacturer from broad external list.",
    },
    {
        "manufacturer_name": "Durand",
        "canonical_match_name": "Durand",
        "manufacturer_type": "boutique-upgrade",
        "coverage_focus": "cottage-boutique-arms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Durand tonearm Talea Tosca geometry effective length overhang",
        "notes": "High-priority missing boutique manufacturer from broad external list.",
    },
    {
        "manufacturer_name": "Goldmund",
        "canonical_match_name": "Goldmund",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Goldmund T3F T5 tonearm geometry effective length overhang",
        "notes": "High-priority missing historic manufacturer from broad external list.",
    },
    {
        "manufacturer_name": "Koshin",
        "canonical_match_name": "Koshin",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Koshin GST-801 GST-1 tonearm geometry effective length overhang",
        "notes": "High-priority missing Japanese arm maker from broad external list.",
    },
    {
        "manufacturer_name": "Micro Seiki",
        "canonical_match_name": "Micro Seiki",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "Micro Seiki MA-505 MA-707 MAX-282 tonearm geometry effective length overhang",
        "notes": "High-priority missing Japanese arm maker from broad external list.",
    },
    {
        "manufacturer_name": "Saec",
        "canonical_match_name": "Saec",
        "manufacturer_type": "standalone-upgrade",
        "coverage_focus": "upgrade-tonearms",
        "priority_tier": "secondary",
        "status": "queued",
        "discovery_source": "external-list-followup",
        "search_terms": "SAEC Saec WE-308SX WE-407/23 WE-8000/ST tonearm geometry effective length overhang",
        "notes": "Existing DB manufacturer matched from the broad external list but not yet promoted into explicit research queue follow-up.",
    },
]


TONEARM_RESEARCH_TARGET_SEEDS: list[dict[str, str]] = [
    {"manufacturer_name": "Tri-Planar", "model_name": "Mk VII", "target_group": "U2", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official site / reviews", "notes": "Current flagship-style upgrade arm."},
    {"manufacturer_name": "Tri-Planar", "model_name": "U2 SE", "target_group": "U2", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official site", "notes": "Current short-arm family target."},
    {"manufacturer_name": "Tri-Planar", "model_name": "U12", "target_group": "U12", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official site", "notes": "Current 12 inch Tri-Planar flagship with official geometry."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official site", "notes": "Replaced by exact Phantom III 9-inch, 10-inch, and 12-inch variants."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III 9-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official page", "notes": "Exact-size target split from generic Phantom III."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III 10-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official page", "notes": "Exact-size target split from generic Phantom III."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III 12-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official page", "notes": "Exact-size target split from generic Phantom III."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III SE", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official site", "notes": "Replaced by exact Phantom III SE 9-inch, 10-inch, and 12-inch variants."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III SE 9-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official page", "notes": "Exact-size Phantom III SE target from official page."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III SE 10-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official page", "notes": "Exact-size Phantom III SE target from official page."},
    {"manufacturer_name": "Graham", "model_name": "Phantom III SE 12-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official page", "notes": "Exact-size Phantom III SE target from official page."},
    {"manufacturer_name": "Graham", "model_name": "Phantom Elite", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official brochure", "notes": "Replaced by exact Phantom Elite 9-inch, 10-inch, and 12-inch variants."},
    {"manufacturer_name": "Graham", "model_name": "Phantom Elite 9-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official brochure", "notes": "Exact-size target split from generic Phantom Elite."},
    {"manufacturer_name": "Graham", "model_name": "Phantom Elite 10-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official brochure", "notes": "Exact-size target split from generic Phantom Elite."},
    {"manufacturer_name": "Graham", "model_name": "Phantom Elite 12-inch", "target_group": "Phantom", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official brochure", "notes": "Exact-size target split from generic Phantom Elite."},
    {"manufacturer_name": "Kuzma", "model_name": "4Point", "target_group": "4Point", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official manuals", "notes": "Replaced by exact 4Point 9-inch, 11-inch, and 14-inch variants."},
    {"manufacturer_name": "Kuzma", "model_name": "4Point 9-inch", "target_group": "4Point", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 9 inch 4Point variant from official page."},
    {"manufacturer_name": "Kuzma", "model_name": "4Point 11-inch", "target_group": "4Point", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 11 inch 4Point variant from official page/manual set."},
    {"manufacturer_name": "Kuzma", "model_name": "4Point 14-inch", "target_group": "4Point", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 14 inch 4Point variant from official page/brochure."},
    {"manufacturer_name": "Kuzma", "model_name": "Stogi S", "target_group": "Stogi", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official manuals", "notes": "Replaced by exact Stogi S 9-inch, 12-inch, and 12-inch VTA variants."},
    {"manufacturer_name": "Kuzma", "model_name": "Stogi S 9-inch", "target_group": "Stogi", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 9 inch Stogi S variant from official page/brochure."},
    {"manufacturer_name": "Kuzma", "model_name": "Air Line", "target_group": "Air Line", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official manuals", "notes": "Replaced by exact Air Line 7-inch variant."},
    {"manufacturer_name": "Kuzma", "model_name": "Air Line 7-inch", "target_group": "Air Line", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact Air Line variant from official page."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official manual", "notes": "Replaced by exact Blackbird 9-inch, 9.3-inch, 10.5-inch, and 12-inch variants."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird 9-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact Blackbird size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird 9.3-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact Blackbird size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird 10.5-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact Blackbird size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird 12-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact Blackbird size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird Farpoint", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official manual", "notes": "Replaced by exact Blackbird Farpoint 9-inch, 9.3-inch, 10.5-inch, and 12-inch variants."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird Farpoint 9-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official manual / site", "notes": "Exact Blackbird Farpoint size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird Farpoint 9.3-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official manual / site", "notes": "Exact Blackbird Farpoint size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird Farpoint 10.5-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official manual / site", "notes": "Exact Blackbird Farpoint size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Blackbird Farpoint 12-inch", "target_group": "Blackbird", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official manual / site", "notes": "Exact Blackbird Farpoint size variant from official Supatrac geometry family."},
    {"manufacturer_name": "Supatrac", "model_name": "Nighthawk", "target_group": "Nighthawk", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "superseded", "source_hint": "official product page", "notes": "Replaced by exact Nighthawk 9-inch, 9.3-inch, 10.5-inch, and 12-inch variants."},
    {"manufacturer_name": "Supatrac", "model_name": "Nighthawk 9-inch", "target_group": "Nighthawk", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 9 inch Nighthawk variant from official page."},
    {"manufacturer_name": "Supatrac", "model_name": "Nighthawk 9.3-inch", "target_group": "Nighthawk", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 9.3 inch Nighthawk variant from official page."},
    {"manufacturer_name": "Supatrac", "model_name": "Nighthawk 10.5-inch", "target_group": "Nighthawk", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 10.5 inch Nighthawk variant from official page."},
    {"manufacturer_name": "Supatrac", "model_name": "Nighthawk 12-inch", "target_group": "Nighthawk", "target_type": "boutique-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official product page", "notes": "Exact 12 inch Nighthawk variant from official page."},
    {"manufacturer_name": "Origin Live", "model_name": "Alliance", "target_group": "9.5 inch", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "install manual", "notes": "Origin Live upgrade family target."},
    {"manufacturer_name": "Origin Live", "model_name": "Onyx", "target_group": "9.5 inch", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "install manual", "notes": "Origin Live upgrade family target."},
    {"manufacturer_name": "Origin Live", "model_name": "Silver", "target_group": "9.5 inch", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "install manual", "notes": "Origin Live upgrade family target."},
    {"manufacturer_name": "Origin Live", "model_name": "Zephyr", "target_group": "9.5 inch", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "install manual", "notes": "Origin Live upgrade family target."},
    {"manufacturer_name": "Rega", "model_name": "RB250", "target_group": "RB", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "manual / spec page", "notes": "Common upgrade arm."},
    {"manufacturer_name": "Rega", "model_name": "RB300", "target_group": "RB", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "manual / spec page", "notes": "Common upgrade arm."},
    {"manufacturer_name": "Rega", "model_name": "RB330", "target_group": "RB", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "official pdf", "notes": "Current production priority."},
    {"manufacturer_name": "VPI", "model_name": "JMW-9", "target_group": "JMW", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "VPI manuals", "notes": "Combo arm family target."},
    {"manufacturer_name": "VPI", "model_name": "JMW-10.5", "target_group": "JMW", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "VPI manuals", "notes": "Combo arm family target."},
    {"manufacturer_name": "VPI", "model_name": "JMW-12", "target_group": "JMW", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "superseded", "source_hint": "VPI manuals", "notes": "Replaced by exact JMW-12 3D target on the current VPI unipivot page."},
    {"manufacturer_name": "VPI", "model_name": "JMW-12 3D", "target_group": "JMW", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official page", "notes": "Exact current 12 inch 3D VPI target."},
    {"manufacturer_name": "Pro-Ject", "model_name": "9cc Evolution", "target_group": "Evolution", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official pdf", "notes": "Public geometry available."},
    {"manufacturer_name": "Pro-Ject", "model_name": "10cc Evolution", "target_group": "Evolution", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official pdf", "notes": "Public geometry available."},
    {"manufacturer_name": "Pro-Ject", "model_name": "12cc Evolution", "target_group": "Evolution", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official pdf", "notes": "Public geometry available."},
    {"manufacturer_name": "Music Hall", "model_name": "MMF-1.3", "target_group": "MMF", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Entry Music Hall turntable with published tonearm geometry."},
    {"manufacturer_name": "Music Hall", "model_name": "MMF-2.3", "target_group": "MMF", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Music Hall turntable with published tonearm geometry."},
    {"manufacturer_name": "Music Hall", "model_name": "MMF-3.3", "target_group": "MMF", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Music Hall turntable with published tonearm geometry."},
    {"manufacturer_name": "Music Hall", "model_name": "MMF-5.3", "target_group": "MMF", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Music Hall turntable with published tonearm geometry."},
    {"manufacturer_name": "Music Hall", "model_name": "MMF-7.3", "target_group": "MMF", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Music Hall turntable with published tonearm geometry."},
    {"manufacturer_name": "Music Hall", "model_name": "MMF-9.3", "target_group": "MMF", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Music Hall turntable with published tonearm geometry."},
    {"manufacturer_name": "TW-Acustic", "model_name": "Raven 10.5", "target_group": "Raven", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official/manualslib", "notes": "Public user manual exists."},
    {"manufacturer_name": "TW-Acustic", "model_name": "Raven 12-inch", "target_group": "Raven", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official follow-up", "notes": "Stereophile 2025 recommended 12 inch Raven arm target."},
    {"manufacturer_name": "Basis Audio", "model_name": "Vector", "target_group": "Vector", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "superseded", "source_hint": "catalog/manual research", "notes": "Replaced by exact Vector 4 variant currently represented in the DB."},
    {"manufacturer_name": "Basis Audio", "model_name": "Vector 4", "target_group": "Vector", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Vector family tonearm with public specs."},
    {"manufacturer_name": "Dr. Feickert", "model_name": "Straight-10", "target_group": "Straight", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "review + dealer listing", "notes": "Current Dr. Feickert branded tonearm with published geometry and mount data."},
    {"manufacturer_name": "Gold Note", "model_name": "B-5.1", "target_group": "B-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "superseded", "source_hint": "official product page", "notes": "Legacy B-5.1 page now resolves to B-5.2 content; replaced by exact B-5.2 target."},
    {"manufacturer_name": "Gold Note", "model_name": "B-5.2", "target_group": "B-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Gold Note arm with public official specs; official legacy B-5.1 URL now serves B-5.2 content."},
    {"manufacturer_name": "Gold Note", "model_name": "B-7 Ceramic", "target_group": "B-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Flagship Gold Note arm with public official specs."},
    {"manufacturer_name": "Ortofon", "model_name": "AS-212R", "target_group": "Reference", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current 9 inch Reference tonearm."},
    {"manufacturer_name": "Ortofon", "model_name": "AS-309R", "target_group": "Reference", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official userguide", "notes": "Current 12 inch Reference tonearm."},
    {"manufacturer_name": "Ortofon", "model_name": "RS-212D", "target_group": "RS D-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current 9 inch dynamic-balance tonearm."},
    {"manufacturer_name": "Ortofon", "model_name": "RS-309D", "target_group": "RS D-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product family", "notes": "Current 12 inch dynamic-balance tonearm."},
    {"manufacturer_name": "Audiomods", "model_name": "Series Six", "target_group": "Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "superseded", "source_hint": "manual/product research", "notes": "Replaced by exact Series Six 9-inch Rega, 9-inch Linn, and 10.5-inch variants."},
    {"manufacturer_name": "Audiomods", "model_name": "Series Six 9-inch Rega", "target_group": "Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official dimensions pdf", "notes": "Exact-model Audiomods Series Six variant from official dimensions sheet."},
    {"manufacturer_name": "Audiomods", "model_name": "Series Six 9-inch Linn", "target_group": "Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official dimensions pdf", "notes": "Exact-model Audiomods Series Six variant from official dimensions sheet."},
    {"manufacturer_name": "Audiomods", "model_name": "Series Six 10.5-inch", "target_group": "Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official dimensions pdf", "notes": "Exact-model Audiomods Series Six variant from official dimensions sheet."},
    {"manufacturer_name": "AMG", "model_name": "9W2", "target_group": "9W", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current 9 inch AMG arm with published geometry."},
    {"manufacturer_name": "AMG", "model_name": "12JT", "target_group": "12J", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current flagship 12 inch AMG arm with published geometry."},
    {"manufacturer_name": "Korf", "model_name": "TA-SF9", "target_group": "TA-SF", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Fixed-headshell Korf reference arm."},
    {"manufacturer_name": "Korf", "model_name": "TA-SF9R", "target_group": "TA-SF", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official product page", "notes": "J-shaped removable-headshell Korf arm listed by Stereophile 2025."},
    {"manufacturer_name": "Korf", "model_name": "TA-SF11R", "target_group": "TA-SF", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Longer removable-headshell Korf arm."},
    {"manufacturer_name": "Schroeder", "model_name": "Reference", "target_group": "Reference", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official mounting scheme", "notes": "Flagship Schroeder pivot arm."},
    {"manufacturer_name": "Schroeder", "model_name": "CB", "target_group": "CB", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "superseded", "source_hint": "manual/product research", "notes": "Replaced by exact CB 9-inch and CB-L 12-inch variants."},
    {"manufacturer_name": "Schroeder", "model_name": "CB 9-inch", "target_group": "CB", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "dealer spec page", "notes": "Exact 9 inch CB variant from detailed spec table."},
    {"manufacturer_name": "Schroeder", "model_name": "CB-L 12-inch", "target_group": "CB", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "dealer spec page", "notes": "Exact 12 inch CB-L variant from detailed spec table."},
    {"manufacturer_name": "Schroeder", "model_name": "Model 2", "target_group": "No.2", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official manual", "notes": "Official manual provides mounting geometry and Baerwald alignment guidance."},
    {"manufacturer_name": "Schroeder", "model_name": "DPS", "target_group": "DPS", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official manual", "notes": "Official manual provides mounting geometry and alignment guidance for the DPS arm."},
    {"manufacturer_name": "Thales", "model_name": "Statement", "target_group": "Reference", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "brochure", "notes": "Official geometry surfaced."},
    {"manufacturer_name": "Thales", "model_name": "Simplicity II", "target_group": "Simplicity", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "brochure", "notes": "Official geometry surfaced."},
    {"manufacturer_name": "SAT", "model_name": "LM-09", "target_group": "LM", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official geometry paper", "notes": "Official geometry surfaced."},
    {"manufacturer_name": "SAT", "model_name": "LM-12", "target_group": "LM", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official geometry paper", "notes": "Official geometry surfaced."},
    {"manufacturer_name": "Acoustic Signature", "model_name": "TA-5000 NEO", "target_group": "TA-5000", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended Acoustic Signature arm."},
    {"manufacturer_name": "Acoustic Signature", "model_name": "TA-7000 NEO 9-inch", "target_group": "TA-7000", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended 9 inch Acoustic Signature flagship arm."},
    {"manufacturer_name": "Acoustic Signature", "model_name": "TA-7000 NEO 12-inch", "target_group": "TA-7000", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended 12 inch Acoustic Signature flagship arm."},
    {"manufacturer_name": "Acoustical Systems", "model_name": "AQUILAR Reference 10-inch", "target_group": "AQUILAR", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 10 inch Acoustical Systems target aligned to current official lineup naming (AQUILAR for 10 inch, AXIOM for 12 inch)."},
    {"manufacturer_name": "Acoustical Systems", "model_name": "AXIOM Reference 12-inch", "target_group": "AXIOM", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended 12 inch AXIOM Reference arm."},
    {"manufacturer_name": "Clearaudio", "model_name": "Tracer Black Carbon Fiber", "target_group": "Tracer", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended Clearaudio arm."},
    {"manufacturer_name": "EMT", "model_name": "909-HI", "target_group": "900-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended EMT 9 inch arm."},
    {"manufacturer_name": "EMT", "model_name": "912", "target_group": "900-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended EMT 12 inch arm."},
    {"manufacturer_name": "J.Sikora", "model_name": "KV12 VTA", "target_group": "KV", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended J.Sikora arm."},
    {"manufacturer_name": "Kuzma", "model_name": "Safir 9", "target_group": "Safir", "target_type": "standalone-upgrade", "priority_tier": "primary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended Kuzma flagship sapphire arm."},
    {"manufacturer_name": "Linn", "model_name": "Ekos SE", "target_group": "Ekos", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended Linn arm."},
    {"manufacturer_name": "Schick", "model_name": "Schick 10.5-inch Tonearm", "target_group": "Schick", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended Thomas Schick 10.5 inch arm."},
    {"manufacturer_name": "Schick", "model_name": "Schick 12-inch Tonearm", "target_group": "Schick", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended Thomas Schick 12 inch arm."},
    {"manufacturer_name": "Sorane", "model_name": "SA-1.2", "target_group": "SA", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official manual / stereophile review", "notes": "Stereophile 2025 recommended Sorane arm already represented in the DB."},
    {"manufacturer_name": "Sorane", "model_name": "TA-1", "target_group": "TA", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile 2025 review / official page", "notes": "Stereophile 2025 listed Sorane TA-1 as a recommendable conventional gimbal-bearing arm."},
    {"manufacturer_name": "ViV Laboratory", "model_name": "Rigid Float HA9", "target_group": "Rigid Float", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended underhung ViV arm."},
    {"manufacturer_name": "VPI", "model_name": "12-inch FatBoy gimbaled", "target_group": "FatBoy", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 recommended VPI gimbaled FatBoy arm."},
    {"manufacturer_name": "Wand", "model_name": "Master 12-inch", "target_group": "Master", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "stereophile review / official page", "notes": "Stereophile 2025 Wand target aligned to current official lineup naming (Plus/Master series; 12-inch Master arm)."},
    {"manufacturer_name": "Vertere Acoustics", "model_name": "Super Groove Precision Tonearm", "target_group": "Super Groove", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official tonearm page", "notes": "Current Vertere SG-1-related precision tonearm page on official site."},
    {"manufacturer_name": "Vertere Acoustics", "model_name": "Super Groove II Pathfinder Tonearm", "target_group": "Super Groove", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official tonearm page", "notes": "Current Vertere SG-II Pathfinder tonearm page on official site."},
    {"manufacturer_name": "Vertere Acoustics", "model_name": "Reference Tonearm (Gen III)", "target_group": "Reference", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official tonearm page", "notes": "Current Vertere flagship reference tonearm page on official site."},
    {"manufacturer_name": "Michell", "model_name": "T2 Tonearm", "target_group": "T-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Michell entry-level tonearm."},
    {"manufacturer_name": "Michell", "model_name": "T8 Tonearm", "target_group": "T-Series", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Michell audiophile tonearm."},
    {"manufacturer_name": "Michell", "model_name": "TecnoArm 2", "target_group": "TecnoArm", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official product page", "notes": "Current Michell reference-class tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-900S", "target_group": "S-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 9-inch S-Type tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-1000S", "target_group": "S-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 10-inch S-Type tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-1200S", "target_group": "S-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 12-inch S-Type tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-12SUS (MH-124S Premium)", "target_group": "Premium", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz premium 12-inch tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-94S", "target_group": "Legacy S-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official end-of-sale page", "notes": "Legacy Glanz 9-inch tonearm still documented on official site."},
    {"manufacturer_name": "Glanz", "model_name": "MH-104S", "target_group": "Legacy S-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official end-of-sale page", "notes": "Legacy Glanz 10-inch tonearm still documented on official site."},
    {"manufacturer_name": "Glanz", "model_name": "MH-124S", "target_group": "Legacy S-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official end-of-sale page", "notes": "Legacy Glanz 12-inch tonearm still documented on official site."},
    {"manufacturer_name": "Glanz", "model_name": "MH-9B", "target_group": "B-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 9-inch B-Type tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-10B", "target_group": "B-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 10-inch B-Type tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-9Bt", "target_group": "Bt-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 9-inch Bt-Type tonearm."},
    {"manufacturer_name": "Glanz", "model_name": "MH-10Bt", "target_group": "Bt-Type", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official collection page", "notes": "Current Glanz 10-inch Bt-Type tonearm."},
    {"manufacturer_name": "Pear Audio", "model_name": "Cornet 1", "target_group": "Cornet", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "catalog / review follow-up", "notes": "Pear Audio tonearm family target from catalog and review references."},
    {"manufacturer_name": "Pear Audio", "model_name": "Cornet 2", "target_group": "Cornet", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "catalog / review follow-up", "notes": "Pear Audio tonearm family target from catalog and review references."},
    {"manufacturer_name": "Pear Audio", "model_name": "Cornet 3", "target_group": "Cornet", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "catalog / review follow-up", "notes": "Pear Audio tonearm family target from catalog and review references."},
    {"manufacturer_name": "Audio Creative", "model_name": "Groovemaster 4", "target_group": "Groovemaster", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "The Ear review / official page", "notes": "Current Groovemaster flagship highlighted by The Ear; available in 9-inch, 10-inch, and 12-inch forms."},
    {"manufacturer_name": "Acoustic Solid", "model_name": "WTB 213", "target_group": "WTB", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official page / distributor specs", "notes": "Acoustic Solid 9-inch class tonearm target from broad manufacturer-gap follow-up."},
    {"manufacturer_name": "Acoustic Solid", "model_name": "WTB 370", "target_group": "WTB", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official page / distributor specs", "notes": "Acoustic Solid premium tonearm target from broad manufacturer-gap follow-up."},
    {"manufacturer_name": "Air Tangent", "model_name": "2A", "target_group": "2-Series", "target_type": "parallel-tracker", "priority_tier": "secondary", "status": "superseded", "source_hint": "manual / archival specs", "notes": "Legacy external-list target retained for traceability but superseded in active follow-up by exact archived official Air Tangent models that are currently corroborated."},
    {"manufacturer_name": "Air Tangent", "model_name": "2B", "target_group": "2-Series", "target_type": "parallel-tracker", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / archival specs", "notes": "Flagship Air Tangent linear tracker target."},
    {"manufacturer_name": "Air Tangent", "model_name": "Reference", "target_group": "Reference", "target_type": "parallel-tracker", "priority_tier": "secondary", "status": "queued", "source_hint": "official archived website", "notes": "Archived official Airtangent homepage exact model preceding the later Model 2002 site revision."},
    {"manufacturer_name": "Air Tangent", "model_name": "Model 2002", "target_group": "2000-Series", "target_type": "parallel-tracker", "priority_tier": "secondary", "status": "queued", "source_hint": "official archived website", "notes": "Archived official Airtangent homepage exact model succeeding the earlier Reference site revision."},
    {"manufacturer_name": "Rabco", "model_name": "SL-8", "target_group": "SL-8", "target_type": "parallel-tracker", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / archival specs", "notes": "Classic Rabco linear-tracking arm target."},
    {"manufacturer_name": "Rabco", "model_name": "SL-8E", "target_group": "SL-8", "target_type": "parallel-tracker", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / archival specs", "notes": "Later Rabco linear-tracking arm target with servo-drive follow-up."},
    {"manufacturer_name": "Alphason Designs", "model_name": "HR-100S", "target_group": "HR-100", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "archival brochure / review", "notes": "Classic Alphason tonearm target."},
    {"manufacturer_name": "Alphason Designs", "model_name": "HR-100MCS", "target_group": "HR-100", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "archival brochure / review", "notes": "Classic Alphason flagship target."},
    {"manufacturer_name": "Audio Note", "model_name": "Arm One/II", "target_group": "Arm", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official page / brochure", "notes": "Current Audio Note entry-level arm target."},
    {"manufacturer_name": "Audio Note", "model_name": "Arm Three/II", "target_group": "Arm", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official page / brochure", "notes": "Current Audio Note higher-tier arm target."},
    {"manufacturer_name": "Durand", "model_name": "Talea", "target_group": "Durand", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official page / review", "notes": "Durand Talea target from broad manufacturer-gap follow-up."},
    {"manufacturer_name": "Durand", "model_name": "Tosca", "target_group": "Durand", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official page / review", "notes": "Durand Tosca target from broad manufacturer-gap follow-up."},
    {"manufacturer_name": "Goldmund", "model_name": "T3F", "target_group": "T-Series", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "archival manual / review", "notes": "Classic Goldmund linear-tracking arm target."},
    {"manufacturer_name": "Goldmund", "model_name": "T5", "target_group": "T-Series", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "archival manual / review", "notes": "Classic Goldmund pivoted-arm target."},
    {"manufacturer_name": "Koshin", "model_name": "GST-801", "target_group": "GST", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / review", "notes": "Classic Koshin tonearm target."},
    {"manufacturer_name": "Koshin", "model_name": "GST-1", "target_group": "GST", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / review", "notes": "Classic Koshin flagship tonearm target."},
    {"manufacturer_name": "Micro Seiki", "model_name": "MA-505", "target_group": "MA", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / archival specs", "notes": "Classic Micro Seiki MA-series arm target."},
    {"manufacturer_name": "Micro Seiki", "model_name": "MA-707", "target_group": "MA", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "manual / archival specs", "notes": "Classic Micro Seiki MA-series arm target."},
    {"manufacturer_name": "Saec", "model_name": "WE-308SX", "target_group": "WE", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "archival specs / review", "notes": "Classic Saec pivoted-arm target promoted from seeded DB presence."},
    {"manufacturer_name": "Saec", "model_name": "WE-407/23", "target_group": "WE", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "archival specs / review", "notes": "Classic Saec 12-inch transcription-arm target promoted from seeded DB presence."},
    {"manufacturer_name": "Dynavector", "model_name": "DV-501/505/507", "target_group": "DV", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Grouped Dynavector arm family currently present only as a legacy DB row."},
    {"manufacturer_name": "Audio-Technica", "model_name": "AT-1001/1005II/1007/1009", "target_group": "AT-1000", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Grouped historic Audio-Technica 10-inch arm family from legacy DB row."},
    {"manufacturer_name": "Audio-Technica", "model_name": "AT-1100/1010/1120", "target_group": "AT-1100", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Grouped historic Audio-Technica arm family from legacy DB row."},
    {"manufacturer_name": "Audio-Technica", "model_name": "AT-1501II", "target_group": "AT-1501", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Historic Audio-Technica exact legacy row to validate against better sources."},
    {"manufacturer_name": "Audio-Technica", "model_name": "AT-1501III, IV", "target_group": "AT-1501", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Grouped historic Audio-Technica AT-1501 variants from legacy DB row."},
    {"manufacturer_name": "Audio-Technica", "model_name": "AT-1503II", "target_group": "AT-1503", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Historic Audio-Technica exact legacy row to validate against better sources."},
    {"manufacturer_name": "Audio-Technica", "model_name": "AT-1503III, IV, IIIa", "target_group": "AT-1503", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Grouped historic Audio-Technica AT-1503 variants from legacy DB row."},
    {"manufacturer_name": "Fidelity Research", "model_name": "FR-12", "target_group": "FR", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Historic Fidelity Research arm family follow-up from legacy DB row."},
    {"manufacturer_name": "Fidelity Research", "model_name": "FR-14/FR-24/FR-54/FR-64", "target_group": "FR", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Grouped Fidelity Research arm family follow-up from legacy DB row."},
    {"manufacturer_name": "Fidelity Research", "model_name": "FR-34", "target_group": "FR", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Historic Fidelity Research exact legacy row to validate against better sources."},
    {"manufacturer_name": "Fidelity Research", "model_name": "FR-66S", "target_group": "FR-66", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Historic Fidelity Research flagship follow-up from legacy DB row."},
    {"manufacturer_name": "Fidelity Research", "model_name": "FR-66fx", "target_group": "FR-66", "target_type": "standalone-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "legacy table / official follow-up", "notes": "Historic Fidelity Research late flagship follow-up from legacy DB row."},
    {"manufacturer_name": "Well Tempered Lab", "model_name": "LTD", "target_group": "LTD", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "superseded", "source_hint": "official manual", "notes": "Replaced by exact LTD 10.5-inch and LTD 16-inch variants."},
    {"manufacturer_name": "Well Tempered Lab", "model_name": "LTD 10.5-inch", "target_group": "LTD", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official manual", "notes": "Hydrated from standalone LTD manual."},
    {"manufacturer_name": "Well Tempered Lab", "model_name": "LTD 16-inch", "target_group": "LTD", "target_type": "boutique-upgrade", "priority_tier": "secondary", "status": "queued", "source_hint": "official brochure", "notes": "Hydrated from Royale 400 brochure."},
    {"manufacturer_name": "TechDAS", "model_name": "Air Force 10", "target_group": "Air Force", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "superseded", "source_hint": "official specs", "notes": "Replaced by exact Air Force 10 10-inch and 12-inch variants."},
    {"manufacturer_name": "TechDAS", "model_name": "Air Force 10 10-inch", "target_group": "Air Force", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official leaflet", "notes": "Hydrated from public TechDAS geometry leaflet."},
    {"manufacturer_name": "TechDAS", "model_name": "Air Force 10 12-inch", "target_group": "Air Force", "target_type": "turntable-combo", "priority_tier": "secondary", "status": "queued", "source_hint": "official leaflet", "notes": "Hydrated from public TechDAS geometry leaflet."},
]


MANUFACTURER_NORMALIZATION_CANDIDATES: list[dict[str, str]] = [
    {
        "legacy_name": "Schroder",
        "canonical_name": "Schroeder",
        "reason": "Legacy transliteration variant; current exact-model enrichment uses 'Schroeder'.",
    },
    {
        "legacy_name": "Wheaton/Tri-Planar",
        "canonical_name": "Tri-Planar",
        "reason": "Legacy predecessor/company label; current exact-model enrichment uses the active 'Tri-Planar' brand.",
    },
    {
        "legacy_name": "J. Sikora",
        "canonical_name": "J.Sikora",
        "reason": "Punctuation variant used in editorial and user-provided references; canonical queue seed uses 'J.Sikora'.",
    },
]


SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS manufacturers (
    manufacturer_id INTEGER PRIMARY KEY,
    name TEXT NOT NULL UNIQUE,
    normalized_name TEXT NOT NULL UNIQUE,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS tonearm_models (
    tonearm_model_id INTEGER PRIMARY KEY,
    manufacturer_id INTEGER NOT NULL REFERENCES manufacturers(manufacturer_id),
    model_name TEXT NOT NULL,
    normalized_model_name TEXT NOT NULL,
    display_name TEXT NOT NULL,
    notes TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(manufacturer_id, normalized_model_name)
);

CREATE TABLE IF NOT EXISTS sources (
    source_id INTEGER PRIMARY KEY,
    source_name TEXT NOT NULL,
    source_url TEXT NOT NULL,
    source_type TEXT NOT NULL,
    trust_level TEXT NOT NULL,
    retrieved_at TEXT NOT NULL,
    local_snapshot_path TEXT,
    notes TEXT,
    UNIQUE(source_url, retrieved_at)
);

CREATE TABLE IF NOT EXISTS ingest_runs (
    ingest_run_id INTEGER PRIMARY KEY,
    started_at TEXT NOT NULL,
    completed_at TEXT,
    status TEXT NOT NULL,
    notes TEXT
);

CREATE TABLE IF NOT EXISTS tonearm_specs (
    tonearm_spec_id INTEGER PRIMARY KEY,
    tonearm_model_id INTEGER NOT NULL REFERENCES tonearm_models(tonearm_model_id),
    source_id INTEGER NOT NULL REFERENCES sources(source_id),
    ingest_run_id INTEGER REFERENCES ingest_runs(ingest_run_id),
    field_name TEXT NOT NULL,
    value_num REAL,
    value_text TEXT,
    unit TEXT,
    raw_value_text TEXT,
    status TEXT NOT NULL,
    confidence REAL,
    notes TEXT,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_tonearm_specs_model_field ON tonearm_specs(tonearm_model_id, field_name);
CREATE INDEX IF NOT EXISTS idx_tonearm_specs_source ON tonearm_specs(source_id);

CREATE TABLE IF NOT EXISTS manufacturer_research_queue (
    manufacturer_queue_id INTEGER PRIMARY KEY,
    manufacturer_name TEXT NOT NULL,
    normalized_name TEXT NOT NULL UNIQUE,
    canonical_match_name TEXT,
    canonical_manufacturer_id INTEGER REFERENCES manufacturers(manufacturer_id),
    manufacturer_type TEXT NOT NULL,
    coverage_focus TEXT NOT NULL,
    priority_tier TEXT NOT NULL,
    status TEXT NOT NULL,
    discovery_source TEXT,
    search_terms TEXT,
    notes TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS tonearm_research_targets (
    target_id INTEGER PRIMARY KEY,
    manufacturer_queue_id INTEGER NOT NULL REFERENCES manufacturer_research_queue(manufacturer_queue_id),
    model_name TEXT NOT NULL,
    normalized_model_name TEXT NOT NULL,
    target_group TEXT,
    target_type TEXT NOT NULL,
    priority_tier TEXT NOT NULL,
    status TEXT NOT NULL,
    source_hint TEXT,
    notes TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE(manufacturer_queue_id, normalized_model_name)
);

CREATE INDEX IF NOT EXISTS idx_manufacturer_research_queue_priority ON manufacturer_research_queue(priority_tier, status);
CREATE INDEX IF NOT EXISTS idx_tonearm_research_targets_queue ON tonearm_research_targets(manufacturer_queue_id, priority_tier, status);

DROP VIEW IF EXISTS manufacturer_research_summary;
DROP VIEW IF EXISTS preferred_tonearm_specs;
CREATE VIEW preferred_tonearm_specs AS
SELECT
    tm.tonearm_model_id,
    m.name AS manufacturer,
    tm.model_name AS model,
    MAX(CASE WHEN ts.field_name = 'effective_length_mm' THEN ts.value_num END) AS effective_length_mm,
    MAX(CASE WHEN ts.field_name = 'overhang_mm' THEN ts.value_num END) AS overhang_mm,
    MAX(CASE WHEN ts.field_name = 'offset_angle_deg' THEN ts.value_num END) AS offset_angle_deg,
    MAX(CASE WHEN ts.field_name = 'null_points' THEN COALESCE(ts.value_text, ts.raw_value_text) END) AS null_points,
    COALESCE(
        MAX(CASE WHEN ts.field_name = 'null_alignment_type' AND COALESCE(ts.value_text, ts.raw_value_text) <> 'Unknown' THEN COALESCE(ts.value_text, ts.raw_value_text) END),
        MAX(CASE WHEN ts.field_name = 'null_alignment_type' THEN COALESCE(ts.value_text, ts.raw_value_text) END),
        'Unknown'
    ) AS null_alignment_type,
    MAX(CASE WHEN ts.field_name = 'effective_mass_g' THEN ts.value_num END) AS effective_mass_g,
    MAX(CASE WHEN ts.field_name = 'cartridge_range_low_g' THEN ts.value_num END) AS cartridge_range_low_g,
    MAX(CASE WHEN ts.field_name = 'cartridge_range_high_g' THEN ts.value_num END) AS cartridge_range_high_g,
    MAX(CASE WHEN ts.field_name = 'arm_mount' THEN COALESCE(ts.value_text, ts.raw_value_text) END) AS arm_mount,
    MAX(CASE WHEN ts.field_name = 'null_point_inner_mm' THEN ts.value_num END) AS null_point_inner_mm,
    MAX(CASE WHEN ts.field_name = 'null_point_outer_mm' THEN ts.value_num END) AS null_point_outer_mm,
    GROUP_CONCAT(DISTINCT s.source_name) AS source_names,
    GROUP_CONCAT(DISTINCT s.source_url) AS source_urls
FROM tonearm_models tm
JOIN manufacturers m ON m.manufacturer_id = tm.manufacturer_id
LEFT JOIN tonearm_specs ts ON ts.tonearm_model_id = tm.tonearm_model_id
LEFT JOIN sources s ON s.source_id = ts.source_id
GROUP BY tm.tonearm_model_id, m.name, tm.model_name;

CREATE VIEW manufacturer_research_summary AS
SELECT
    mrq.manufacturer_queue_id,
    mrq.manufacturer_name,
    mrq.canonical_match_name,
    mrq.manufacturer_type,
    mrq.coverage_focus,
    mrq.priority_tier,
    mrq.status,
    mrq.discovery_source,
    mrq.search_terms,
    mrq.notes,
    m.name AS canonical_db_manufacturer,
    COUNT(DISTINCT tm.tonearm_model_id) AS current_db_model_count,
    COUNT(DISTINCT trt.target_id) AS target_model_count
FROM manufacturer_research_queue mrq
LEFT JOIN manufacturers m ON m.manufacturer_id = mrq.canonical_manufacturer_id
LEFT JOIN tonearm_models tm ON tm.manufacturer_id = mrq.canonical_manufacturer_id
LEFT JOIN tonearm_research_targets trt ON trt.manufacturer_queue_id = mrq.manufacturer_queue_id
GROUP BY
    mrq.manufacturer_queue_id,
    mrq.manufacturer_name,
    mrq.canonical_match_name,
    mrq.manufacturer_type,
    mrq.coverage_focus,
    mrq.priority_tier,
    mrq.status,
    mrq.discovery_source,
    mrq.search_terms,
    mrq.notes,
    m.name;
"""


@dataclass(frozen=True)
class SourceSnapshot:
    source_name: str
    source_url: str
    source_type: str
    trust_level: str
    retrieved_at: str
    local_snapshot_path: str
    notes: str | None = None


@dataclass(frozen=True)
class ParsedTonearmRow:
    manufacturer: str
    model: str
    effective_length_mm: float | None
    overhang_mm: float | None
    offset_angle_deg: float | None
    null_points: str | None
    null_point_inner_mm: float | None
    null_point_outer_mm: float | None
    notes: str | None
    raw_effective_length: str
    raw_overhang: str
    raw_offset_angle: str
    raw_null_points: str


def build_tonearm_database(base_dir: str | Path) -> dict[str, Path]:
    base = Path(base_dir)
    source_dir = base / "sources"
    export_dir = base / "exports"
    source_dir.mkdir(parents=True, exist_ok=True)
    export_dir.mkdir(parents=True, exist_ok=True)

    retrieved_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    html = requests.get(BIGLOBE_ARMDATA_URL, timeout=60)
    html.raise_for_status()
    snapshot_path = source_dir / "biglobe_armdata.html"
    snapshot_path.write_text(html.text, encoding="cp1252", errors="ignore")

    snapshot = SourceSnapshot(
        source_name="BIGLOBE Arm Data",
        source_url=BIGLOBE_ARMDATA_URL,
        source_type="secondary_table",
        trust_level="secondary",
        retrieved_at=retrieved_at,
        local_snapshot_path=str(snapshot_path.relative_to(base)),
        notes="Public arm geometry table. Effective mass/cartridge range/arm mount mostly absent.",
    )

    rows = parse_biglobe_armdata(html.text)

    db_path = base / "tonearms.db"
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(SCHEMA_SQL)
        ingest_run_id = _insert_ingest_run(conn, started_at=retrieved_at, notes="Initial hydration from BIGLOBE Arm Data")
        source_id = _insert_source(conn, snapshot)
        for row in rows:
            model_id = _upsert_model(conn, row.manufacturer, row.model, row.notes)
            _insert_spec(conn, model_id, source_id, ingest_run_id, "effective_length_mm", row.effective_length_mm, "mm", row.raw_effective_length)
            _insert_spec(conn, model_id, source_id, ingest_run_id, "overhang_mm", row.overhang_mm, "mm", row.raw_overhang)
            _insert_spec(conn, model_id, source_id, ingest_run_id, "offset_angle_deg", row.offset_angle_deg, "deg", row.raw_offset_angle)
            _insert_spec(conn, model_id, source_id, ingest_run_id, "null_points", None, None, row.raw_null_points, value_text=row.null_points)
            _insert_spec(conn, model_id, source_id, ingest_run_id, "null_point_inner_mm", row.null_point_inner_mm, "mm", row.raw_null_points)
            _insert_spec(conn, model_id, source_id, ingest_run_id, "null_point_outer_mm", row.null_point_outer_mm, "mm", row.raw_null_points)
        _ensure_null_alignment_type_defaults(conn, ingest_run_id)
        _complete_ingest_run(conn, ingest_run_id)
        _sync_research_queue(conn)
        conn.commit()
        outputs = export_database(conn, export_dir)
    finally:
        conn.close()

    schema_path = base / "schema.sql"
    schema_path.write_text(SCHEMA_SQL.strip() + "\n", encoding="utf-8")
    readme_path = base / "README.md"
    readme_path.write_text(_build_readme(rows, snapshot), encoding="utf-8")

    outputs.update(
        {
            "db_path": db_path,
            "schema_sql": schema_path,
            "source_snapshot": snapshot_path,
            "readme": readme_path,
        }
    )
    return outputs


def enrich_tonearm_database(base_dir: str | Path, enrichment_csv: str | Path) -> dict[str, Path]:
    base = Path(base_dir)
    db_path = base / "tonearms.db"
    export_dir = base / "exports"
    if not db_path.exists():
        raise FileNotFoundError(f"tonearm database does not exist: {db_path}")

    conn = sqlite3.connect(db_path)
    try:
        conn.executescript("DROP VIEW IF EXISTS manufacturer_research_summary; DROP VIEW IF EXISTS preferred_tonearm_specs;")
        conn.executescript(SCHEMA_SQL)
        started_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        ingest_run_id = _insert_ingest_run(conn, started_at=started_at, notes=f"Enrichment import from {Path(enrichment_csv).name}")
        source_cache: dict[tuple[str, str], int] = {}
        with Path(enrichment_csv).open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                source_key = (row["source_name"], row["source_url"])
                if source_key not in source_cache:
                    cursor = conn.execute(
                        """
                        INSERT INTO sources(source_name, source_url, source_type, trust_level, retrieved_at, local_snapshot_path, notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            row["source_name"],
                            row["source_url"],
                            row.get("source_type") or "manual",
                            row.get("trust_level") or "manual",
                            started_at,
                            row.get("local_snapshot_path") or None,
                            row.get("source_notes") or None,
                        ),
                    )
                    source_cache[source_key] = int(cursor.lastrowid)
                model_id = _upsert_model(conn, row["manufacturer"], row["model"], row.get("model_notes") or None)
                value_num_text = (row.get("value_num") or "").strip()
                value_num = float(value_num_text) if value_num_text else None
                confidence_text = (row.get("confidence") or "").strip()
                confidence = float(confidence_text) if confidence_text else 0.8
                field_name = row["field_name"]
                value_text = row.get("value_text") or None
                raw_value_text = row.get("raw_value_text") or None
                if field_name == "null_alignment_type":
                    normalized_alignment_type = normalize_null_alignment_type(value_text or raw_value_text)
                    value_text = normalized_alignment_type or "Unknown"
                    raw_value_text = raw_value_text or value_text
                    value_num = None
                _insert_spec(
                    conn,
                    model_id,
                    source_cache[source_key],
                    ingest_run_id,
                    field_name,
                    value_num,
                    row.get("unit") or None,
                    raw_value_text,
                    value_text=value_text,
                    status=row.get("status") or ("parsed" if value_num is not None or value_text else "raw_only"),
                    confidence=confidence,
                    notes=row.get("field_notes") or None,
                )
        _ensure_null_alignment_type_defaults(conn, ingest_run_id)
        _complete_ingest_run(conn, ingest_run_id)
        _sync_research_queue(conn)
        conn.commit()
        return export_database(conn, export_dir)
    finally:
        conn.close()


def sync_tonearm_research_queue(base_dir: str | Path) -> dict[str, Path]:
    base = Path(base_dir)
    db_path = base / "tonearms.db"
    export_dir = base / "exports"
    if not db_path.exists():
        raise FileNotFoundError(f"tonearm database does not exist: {db_path}")

    conn = sqlite3.connect(db_path)
    try:
        conn.executescript("DROP VIEW IF EXISTS manufacturer_research_summary; DROP VIEW IF EXISTS preferred_tonearm_specs;")
        conn.executescript(SCHEMA_SQL)
        _apply_manufacturer_normalization_candidates(conn)
        _sync_research_queue(conn)
        _ensure_null_alignment_type_defaults(conn)
        conn.commit()
        return export_database(conn, export_dir)
    finally:
        conn.close()


def normalize_tonearm_manufacturers(base_dir: str | Path) -> dict[str, Path]:
    base = Path(base_dir)
    db_path = base / "tonearms.db"
    export_dir = base / "exports"
    if not db_path.exists():
        raise FileNotFoundError(f"tonearm database does not exist: {db_path}")

    conn = sqlite3.connect(db_path)
    try:
        conn.executescript("DROP VIEW IF EXISTS manufacturer_research_summary; DROP VIEW IF EXISTS preferred_tonearm_specs;")
        conn.executescript(SCHEMA_SQL)
        _apply_manufacturer_normalization_candidates(conn)
        _sync_research_queue(conn)
        _ensure_null_alignment_type_defaults(conn)
        conn.commit()
        return export_database(conn, export_dir)
    finally:
        conn.close()


def import_tonearm_workbook_gaps(
    base_dir: str | Path,
    workbook_path: str | Path,
) -> dict[str, Path | int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency availability varies by environment
        raise RuntimeError("openpyxl is required to import tonearm workbook gaps") from exc

    base = Path(base_dir)
    db_path = base / "tonearms.db"
    export_dir = base / "exports"
    if not db_path.exists():
        raise FileNotFoundError(f"tonearm database does not exist: {db_path}")

    workbook = Path(workbook_path)
    if not workbook.exists():
        raise FileNotFoundError(f"tonearm workbook does not exist: {workbook}")

    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "Main" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"tonearm workbook missing required 'Main' sheet: {workbook}")

    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    report_rows: list[dict[str, str]] = []
    imported_target_count = 0
    imported_manufacturer_count = 0

    conn = sqlite3.connect(db_path)
    try:
        conn.executescript("DROP VIEW IF EXISTS manufacturer_research_summary; DROP VIEW IF EXISTS preferred_tonearm_specs;")
        conn.executescript(SCHEMA_SQL)
        _apply_manufacturer_normalization_candidates(conn)

        existing_model_keys = {
            (normalize_name(manufacturer), normalize_name(model_name))
            for manufacturer, model_name in conn.execute(
                """
                SELECT m.name, tm.model_name
                FROM tonearm_models tm
                JOIN manufacturers m ON m.manufacturer_id = tm.manufacturer_id
                """
            ).fetchall()
        }
        existing_target_keys = {
            (normalize_name(manufacturer_name), normalize_name(model_name))
            for manufacturer_name, model_name in conn.execute(
                """
                SELECT mrq.manufacturer_name, trt.model_name
                FROM tonearm_research_targets trt
                JOIN manufacturer_research_queue mrq ON mrq.manufacturer_queue_id = trt.manufacturer_queue_id
                """
            ).fetchall()
        }
        seen_workbook_keys: set[tuple[str, str]] = set()

        for row in wb["Main"].iter_rows(min_row=2, values_only=True):
            manufacturer = _coerce_workbook_cell_text(row[0])
            model = _coerce_workbook_cell_text(row[1])
            if not manufacturer or not model:
                continue

            canonical_manufacturer = _canonicalize_workbook_manufacturer(manufacturer)
            normalized_key = (normalize_name(canonical_manufacturer), normalize_name(model))
            effective_length = _coerce_workbook_cell_text(row[3] if len(row) > 3 else None) or _coerce_workbook_cell_text(row[2] if len(row) > 2 else None)
            effective_length_source = _coerce_workbook_cell_text(row[4] if len(row) > 4 else None)
            mounting_distance = _coerce_workbook_cell_text(row[5] if len(row) > 5 else None)
            mounting_distance_source = _coerce_workbook_cell_text(row[6] if len(row) > 6 else None)
            workbook_notes = _coerce_workbook_cell_text(row[7] if len(row) > 7 else None)

            report_row = {
                "source_manufacturer": manufacturer,
                "canonical_manufacturer": canonical_manufacturer,
                "model_name": model,
                "effective_length_raw": effective_length,
                "effective_length_source": effective_length_source,
                "mounting_distance_raw": mounting_distance,
                "mounting_distance_source": mounting_distance_source,
                "workbook_notes": workbook_notes,
                "action": "",
                "reason": "",
            }

            if _is_probable_non_tonearm_workbook_row(canonical_manufacturer, model):
                _delete_workbook_target_if_present(
                    conn,
                    workbook_name=workbook.name,
                    manufacturer_name=canonical_manufacturer,
                    model_name=model,
                )
                report_row["action"] = "skipped"
                report_row["reason"] = "filtered_non_tonearm_row"
                report_rows.append(report_row)
                continue
            if normalized_key in seen_workbook_keys:
                report_row["action"] = "skipped"
                report_row["reason"] = "duplicate_in_workbook"
                report_rows.append(report_row)
                continue
            seen_workbook_keys.add(normalized_key)
            if normalized_key in existing_model_keys:
                report_row["action"] = "skipped"
                report_row["reason"] = "already_in_tonearm_models"
                report_rows.append(report_row)
                continue
            if normalized_key in existing_target_keys:
                report_row["action"] = "skipped"
                report_row["reason"] = "already_in_research_targets"
                report_rows.append(report_row)
                continue

            manufacturer_exists = conn.execute(
                "SELECT 1 FROM manufacturer_research_queue WHERE normalized_name = ?",
                (normalize_name(canonical_manufacturer),),
            ).fetchone()
            if manufacturer_exists is None:
                _upsert_manufacturer_research_seed(
                    conn,
                    {
                        "manufacturer_name": canonical_manufacturer,
                        "canonical_match_name": canonical_manufacturer,
                        "manufacturer_type": "external-reference",
                        "coverage_focus": "workbook-gap-import",
                        "priority_tier": "backlog",
                        "status": "queued",
                        "discovery_source": f"user-workbook:{workbook.name}",
                        "search_terms": f"{canonical_manufacturer} tonearm effective length pivot to spindle official manual specs",
                        "notes": f"Queued from user workbook import: {workbook.name}",
                    },
                    now,
                )
                imported_manufacturer_count += 1

            _upsert_tonearm_research_target_seed(
                conn,
                {
                    "manufacturer_name": canonical_manufacturer,
                    "model_name": model,
                    "target_group": model,
                    "target_type": "tonearm",
                    "priority_tier": "backlog",
                    "status": "queued",
                    "source_hint": f"user workbook: {workbook.name}",
                    "notes": _build_workbook_gap_notes(
                        workbook_name=workbook.name,
                        source_manufacturer=manufacturer,
                        effective_length=effective_length,
                        effective_length_source=effective_length_source,
                        mounting_distance=mounting_distance,
                        mounting_distance_source=mounting_distance_source,
                        workbook_notes=workbook_notes,
                    ),
                },
                now,
            )
            existing_target_keys.add(normalized_key)
            imported_target_count += 1
            report_row["action"] = "imported"
            report_row["reason"] = "added_research_target"
            report_rows.append(report_row)

        _sync_research_queue(conn)
        _ensure_null_alignment_type_defaults(conn)
        conn.commit()
        outputs = export_database(conn, export_dir)
    finally:
        conn.close()
        wb.close()

    report_path = export_dir / "tonearm_workbook_gap_report.csv"
    with report_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "source_manufacturer",
                "canonical_manufacturer",
                "model_name",
                "effective_length_raw",
                "effective_length_source",
                "mounting_distance_raw",
                "mounting_distance_source",
                "workbook_notes",
                "action",
                "reason",
            ],
        )
        writer.writeheader()
        writer.writerows(report_rows)

    outputs["tonearm_workbook_gap_report_csv"] = report_path
    outputs["imported_target_count"] = imported_target_count
    outputs["imported_manufacturer_count"] = imported_manufacturer_count
    return outputs


def _sync_research_queue(conn: sqlite3.Connection) -> None:
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    existing_manufacturers = conn.execute(
        "SELECT manufacturer_id, name FROM manufacturers ORDER BY name"
    ).fetchall()
    for manufacturer_id, name in existing_manufacturers:
        normalized_name = normalize_name(name)
        existing = conn.execute(
            "SELECT manufacturer_queue_id, created_at FROM manufacturer_research_queue WHERE normalized_name = ?",
            (normalized_name,),
        ).fetchone()
        if existing is None:
            conn.execute(
                """
                INSERT INTO manufacturer_research_queue(
                    manufacturer_name, normalized_name, canonical_match_name, canonical_manufacturer_id,
                    manufacturer_type, coverage_focus, priority_tier, status, discovery_source,
                    search_terms, notes, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    name,
                    normalized_name,
                    name,
                    int(manufacturer_id),
                    "legacy-import",
                    "existing-db",
                    "backlog",
                    "seeded-from-db",
                    "auto-seeded",
                    None,
                    "Auto-seeded from current manufacturer table.",
                    now,
                    now,
                ),
            )
        else:
            conn.execute(
                """
                UPDATE manufacturer_research_queue
                SET canonical_manufacturer_id = ?,
                    canonical_match_name = COALESCE(canonical_match_name, ?),
                    updated_at = ?
                WHERE manufacturer_queue_id = ?
                """,
                (int(manufacturer_id), name, now, int(existing[0])),
            )

    for seed in MANUFACTURER_RESEARCH_SEEDS:
        _upsert_manufacturer_research_seed(conn, seed, now)

    for seed in TONEARM_RESEARCH_TARGET_SEEDS:
        _upsert_tonearm_research_target_seed(conn, seed, now)

    _refresh_target_hydration_statuses(conn, now)
    _refresh_manufacturer_hydration_statuses(conn, now)


NON_DERIVED_TARGET_STATUSES = {"superseded"}


def _apply_manufacturer_normalization_candidates(conn: sqlite3.Connection) -> None:
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    for candidate in MANUFACTURER_NORMALIZATION_CANDIDATES:
        _merge_manufacturer_into_canonical(
            conn,
            legacy_name=candidate["legacy_name"],
            canonical_name=candidate["canonical_name"],
            now=now,
        )


def _merge_manufacturer_into_canonical(
    conn: sqlite3.Connection,
    *,
    legacy_name: str,
    canonical_name: str,
    now: str,
) -> None:
    legacy_row = conn.execute(
        "SELECT manufacturer_id FROM manufacturers WHERE name = ?",
        (legacy_name,),
    ).fetchone()
    canonical_row = conn.execute(
        "SELECT manufacturer_id FROM manufacturers WHERE name = ?",
        (canonical_name,),
    ).fetchone()
    if legacy_row is None or canonical_row is None:
        return

    legacy_manufacturer_id = int(legacy_row[0])
    canonical_manufacturer_id = int(canonical_row[0])
    if legacy_manufacturer_id == canonical_manufacturer_id:
        return

    legacy_models = conn.execute(
        """
        SELECT tonearm_model_id, model_name, normalized_model_name
        FROM tonearm_models
        WHERE manufacturer_id = ?
        ORDER BY tonearm_model_id
        """,
        (legacy_manufacturer_id,),
    ).fetchall()
    for tonearm_model_id, model_name, normalized_model_name in legacy_models:
        canonical_model_row = conn.execute(
            """
            SELECT tonearm_model_id
            FROM tonearm_models
            WHERE manufacturer_id = ? AND normalized_model_name = ?
            """,
            (canonical_manufacturer_id, normalized_model_name),
        ).fetchone()
        if canonical_model_row is None:
            conn.execute(
                """
                UPDATE tonearm_models
                SET manufacturer_id = ?,
                    display_name = ?
                WHERE tonearm_model_id = ?
                """,
                (canonical_manufacturer_id, f"{canonical_name} {model_name}", int(tonearm_model_id)),
            )
            continue

        canonical_model_id = int(canonical_model_row[0])
        conn.execute(
            "UPDATE tonearm_specs SET tonearm_model_id = ? WHERE tonearm_model_id = ?",
            (canonical_model_id, int(tonearm_model_id)),
        )
        conn.execute(
            "DELETE FROM tonearm_models WHERE tonearm_model_id = ?",
            (int(tonearm_model_id),),
        )

    conn.execute(
        """
        UPDATE manufacturer_research_queue
        SET canonical_manufacturer_id = ?,
            canonical_match_name = CASE
                WHEN canonical_match_name = ? THEN ?
                ELSE canonical_match_name
            END,
            updated_at = ?
        WHERE canonical_manufacturer_id = ?
        """,
        (canonical_manufacturer_id, legacy_name, canonical_name, now, legacy_manufacturer_id),
    )

    legacy_queue_row = conn.execute(
        "SELECT manufacturer_queue_id FROM manufacturer_research_queue WHERE normalized_name = ?",
        (normalize_name(legacy_name),),
    ).fetchone()
    canonical_queue_row = conn.execute(
        "SELECT manufacturer_queue_id FROM manufacturer_research_queue WHERE normalized_name = ?",
        (normalize_name(canonical_name),),
    ).fetchone()
    if legacy_queue_row is not None and canonical_queue_row is not None:
        legacy_queue_id = int(legacy_queue_row[0])
        canonical_queue_id = int(canonical_queue_row[0])
        legacy_targets = conn.execute(
            "SELECT target_id, normalized_model_name FROM tonearm_research_targets WHERE manufacturer_queue_id = ?",
            (legacy_queue_id,),
        ).fetchall()
        for target_id, normalized_model_name in legacy_targets:
            existing_target_row = conn.execute(
                """
                SELECT target_id
                FROM tonearm_research_targets
                WHERE manufacturer_queue_id = ? AND normalized_model_name = ?
                """,
                (canonical_queue_id, normalized_model_name),
            ).fetchone()
            if existing_target_row is None:
                conn.execute(
                    "UPDATE tonearm_research_targets SET manufacturer_queue_id = ?, updated_at = ? WHERE target_id = ?",
                    (canonical_queue_id, now, int(target_id)),
                )
            else:
                conn.execute(
                    "DELETE FROM tonearm_research_targets WHERE target_id = ?",
                    (int(target_id),),
                )
        conn.execute(
            "DELETE FROM manufacturer_research_queue WHERE manufacturer_queue_id = ?",
            (legacy_queue_id,),
        )

    conn.execute(
        "DELETE FROM manufacturers WHERE manufacturer_id = ?",
        (legacy_manufacturer_id,),
    )


def _refresh_target_hydration_statuses(conn: sqlite3.Connection, now: str) -> None:
    target_rows = conn.execute(
        """
        SELECT
            trt.target_id,
            trt.status,
            trt.normalized_model_name,
            mrq.canonical_manufacturer_id
        FROM tonearm_research_targets trt
        JOIN manufacturer_research_queue mrq ON mrq.manufacturer_queue_id = trt.manufacturer_queue_id
        ORDER BY trt.target_id
        """
    ).fetchall()
    for target_id, status, normalized_model_name, canonical_manufacturer_id in target_rows:
        if status in NON_DERIVED_TARGET_STATUSES or canonical_manufacturer_id is None:
            continue
        matched = conn.execute(
            """
            SELECT 1
            FROM tonearm_models
            WHERE manufacturer_id = ? AND normalized_model_name = ?
            LIMIT 1
            """,
            (int(canonical_manufacturer_id), normalized_model_name),
        ).fetchone()
        derived_status = "hydrated" if matched is not None else "queued"
        if derived_status != status:
            conn.execute(
                "UPDATE tonearm_research_targets SET status = ?, updated_at = ? WHERE target_id = ?",
                (derived_status, now, int(target_id)),
            )


def _refresh_manufacturer_hydration_statuses(conn: sqlite3.Connection, now: str) -> None:
    queue_rows = conn.execute(
        """
        SELECT manufacturer_queue_id, priority_tier, status
        FROM manufacturer_research_queue
        ORDER BY manufacturer_queue_id
        """
    ).fetchall()
    for manufacturer_queue_id, priority_tier, status in queue_rows:
        if priority_tier == "backlog":
            continue
        counts = conn.execute(
            """
            SELECT
                COUNT(*) AS active_targets,
                SUM(CASE WHEN status = 'hydrated' THEN 1 ELSE 0 END) AS hydrated_targets
            FROM tonearm_research_targets
            WHERE manufacturer_queue_id = ?
              AND status IN ('queued', 'hydrated')
            """,
            (int(manufacturer_queue_id),),
        ).fetchone()
        if counts is None:
            continue
        active_targets = int(counts[0] or 0)
        hydrated_targets = int(counts[1] or 0)
        if active_targets == 0:
            continue
        if hydrated_targets == 0:
            derived_status = "queued"
        elif hydrated_targets == active_targets:
            derived_status = "hydrated"
        else:
            derived_status = "hydrated-partial"
        if derived_status != status:
            conn.execute(
                "UPDATE manufacturer_research_queue SET status = ?, updated_at = ? WHERE manufacturer_queue_id = ?",
                (derived_status, now, int(manufacturer_queue_id)),
            )


def export_database(conn: sqlite3.Connection, export_dir: Path) -> dict[str, Path]:
    outputs: dict[str, Path] = {}
    outputs["manufacturers_csv"] = _export_query(
        conn,
        "SELECT manufacturer_id, name, normalized_name, created_at FROM manufacturers ORDER BY name",
        export_dir / "manufacturers.csv",
    )
    outputs["tonearm_models_csv"] = _export_query(
        conn,
        """
        SELECT tm.tonearm_model_id, m.name AS manufacturer, tm.model_name, tm.normalized_model_name, tm.display_name, tm.notes, tm.created_at
        FROM tonearm_models tm
        JOIN manufacturers m ON m.manufacturer_id = tm.manufacturer_id
        ORDER BY m.name, tm.model_name
        """,
        export_dir / "tonearm_models.csv",
    )
    outputs["sources_csv"] = _export_query(
        conn,
        "SELECT source_id, source_name, source_url, source_type, trust_level, retrieved_at, local_snapshot_path, notes FROM sources ORDER BY source_id",
        export_dir / "sources.csv",
    )
    outputs["tonearm_specs_csv"] = _export_query(
        conn,
        """
        SELECT ts.tonearm_spec_id, m.name AS manufacturer, tm.model_name, ts.field_name, ts.value_num, ts.value_text, ts.unit, ts.raw_value_text, ts.status, ts.confidence, ts.notes, s.source_name, s.source_url
        FROM tonearm_specs ts
        JOIN tonearm_models tm ON tm.tonearm_model_id = ts.tonearm_model_id
        JOIN manufacturers m ON m.manufacturer_id = tm.manufacturer_id
        JOIN sources s ON s.source_id = ts.source_id
        ORDER BY m.name, tm.model_name, ts.field_name
        """,
        export_dir / "tonearm_specs.csv",
    )
    outputs["tonearms_csv"] = _export_query(
        conn,
        """
        SELECT
            manufacturer AS Manufacturer,
            model AS Model,
            effective_length_mm AS "Effective Length",
            overhang_mm AS Overhang,
            offset_angle_deg AS "Offset Angle",
            null_points AS "Null Points",
            null_alignment_type AS "Null Alignment Type",
            effective_mass_g AS "Effective Mass",
            cartridge_range_low_g AS "Cartridge Range Low",
            cartridge_range_high_g AS "Cartridge Range High",
            arm_mount AS "Arm Mount",
            null_point_inner_mm AS "Null Point Inner",
            null_point_outer_mm AS "Null Point Outer",
            source_names AS Sources
        FROM preferred_tonearm_specs
        ORDER BY manufacturer, model
        """,
        export_dir / "tonearms.csv",
    )
    outputs["manufacturer_research_queue_csv"] = _export_query(
        conn,
        """
        SELECT
            manufacturer_name,
            canonical_match_name,
            manufacturer_type,
            coverage_focus,
            priority_tier,
            status,
            discovery_source,
            search_terms,
            notes
        FROM manufacturer_research_queue
        ORDER BY
            CASE priority_tier WHEN 'primary' THEN 1 WHEN 'secondary' THEN 2 ELSE 3 END,
            manufacturer_name
        """,
        export_dir / "manufacturer_research_queue.csv",
    )
    outputs["tonearm_research_targets_csv"] = _export_query(
        conn,
        """
        SELECT
            mrq.manufacturer_name,
            trt.model_name,
            trt.target_group,
            trt.target_type,
            trt.priority_tier,
            trt.status,
            trt.source_hint,
            trt.notes
        FROM tonearm_research_targets trt
        JOIN manufacturer_research_queue mrq ON mrq.manufacturer_queue_id = trt.manufacturer_queue_id
        ORDER BY
            CASE trt.priority_tier WHEN 'primary' THEN 1 WHEN 'secondary' THEN 2 ELSE 3 END,
            mrq.manufacturer_name,
            trt.model_name
        """,
        export_dir / "tonearm_research_targets.csv",
    )
    outputs["manufacturer_research_summary_csv"] = _export_query(
        conn,
        """
        SELECT
            manufacturer_name,
            canonical_db_manufacturer,
            manufacturer_type,
            coverage_focus,
            priority_tier,
            status,
            current_db_model_count,
            target_model_count,
            discovery_source,
            search_terms,
            notes
        FROM manufacturer_research_summary
        ORDER BY
            CASE priority_tier WHEN 'primary' THEN 1 WHEN 'secondary' THEN 2 ELSE 3 END,
            manufacturer_name
        """,
        export_dir / "manufacturer_research_summary.csv",
    )
    outputs["model_source_audit_csv"] = _export_query(
        conn,
        """
        SELECT
            m.name AS manufacturer,
            tm.model_name,
            COUNT(DISTINCT s.source_id) AS source_count,
            GROUP_CONCAT(DISTINCT s.trust_level) AS trust_levels,
            GROUP_CONCAT(DISTINCT s.source_type) AS source_types,
            MAX(CASE WHEN s.trust_level = 'official' THEN 1 ELSE 0 END) AS has_official_source,
            MAX(CASE WHEN s.trust_level = 'secondary' THEN 1 ELSE 0 END) AS has_secondary_source,
            MAX(CASE WHEN s.source_type = 'dealer_page' THEN 1 ELSE 0 END) AS has_dealer_source,
            MAX(CASE WHEN s.source_type = 'review' THEN 1 ELSE 0 END) AS has_review_source,
            MAX(CASE WHEN s.source_type = 'manual_mirror' THEN 1 ELSE 0 END) AS has_manual_mirror_source,
            CASE
                WHEN MAX(CASE WHEN s.trust_level = 'official' THEN 1 ELSE 0 END) = 0
                 AND MAX(CASE WHEN s.trust_level = 'secondary' THEN 1 ELSE 0 END) = 1
                THEN 1 ELSE 0
            END AS relies_on_secondary_only,
            CASE
                WHEN MAX(CASE WHEN s.trust_level = 'official' THEN 1 ELSE 0 END) = 0
                THEN 1 ELSE 0
            END AS needs_official_followup
        FROM tonearm_models tm
        JOIN manufacturers m ON m.manufacturer_id = tm.manufacturer_id
        LEFT JOIN tonearm_specs ts ON ts.tonearm_model_id = tm.tonearm_model_id
        LEFT JOIN sources s ON s.source_id = ts.source_id
        GROUP BY tm.tonearm_model_id, m.name, tm.model_name
        ORDER BY m.name, tm.model_name
        """,
        export_dir / "model_source_audit.csv",
    )
    outputs["models_needing_source_upgrade_csv"] = _export_query(
        conn,
        """
        SELECT
            manufacturer,
            model_name,
            source_count,
            trust_levels,
            source_types,
            has_dealer_source,
            has_review_source,
            has_manual_mirror_source,
            relies_on_secondary_only,
            needs_official_followup
        FROM (
            SELECT
                m.name AS manufacturer,
                tm.model_name,
                COUNT(DISTINCT s.source_id) AS source_count,
                GROUP_CONCAT(DISTINCT s.trust_level) AS trust_levels,
                GROUP_CONCAT(DISTINCT s.source_type) AS source_types,
                MAX(CASE WHEN s.trust_level = 'official' THEN 1 ELSE 0 END) AS has_official_source,
                MAX(CASE WHEN s.source_type = 'dealer_page' THEN 1 ELSE 0 END) AS has_dealer_source,
                MAX(CASE WHEN s.source_type = 'review' THEN 1 ELSE 0 END) AS has_review_source,
                MAX(CASE WHEN s.source_type = 'manual_mirror' THEN 1 ELSE 0 END) AS has_manual_mirror_source,
                CASE
                    WHEN MAX(CASE WHEN s.trust_level = 'official' THEN 1 ELSE 0 END) = 0
                     AND MAX(CASE WHEN s.trust_level = 'secondary' THEN 1 ELSE 0 END) = 1
                    THEN 1 ELSE 0
                END AS relies_on_secondary_only,
                CASE
                    WHEN MAX(CASE WHEN s.trust_level = 'official' THEN 1 ELSE 0 END) = 0
                    THEN 1 ELSE 0
                END AS needs_official_followup
            FROM tonearm_models tm
            JOIN manufacturers m ON m.manufacturer_id = tm.manufacturer_id
            LEFT JOIN tonearm_specs ts ON ts.tonearm_model_id = tm.tonearm_model_id
            LEFT JOIN sources s ON s.source_id = ts.source_id
            GROUP BY tm.tonearm_model_id, m.name, tm.model_name
        )
        WHERE needs_official_followup = 1
        ORDER BY manufacturer, model_name
        """,
        export_dir / "models_needing_source_upgrade.csv",
    )

    normalization_rows: list[tuple[object, ...]] = []
    for candidate in MANUFACTURER_NORMALIZATION_CANDIDATES:
        legacy_row = conn.execute(
            "SELECT manufacturer_id, normalized_name FROM manufacturers WHERE name = ?",
            (candidate["legacy_name"],),
        ).fetchone()
        canonical_row = conn.execute(
            "SELECT manufacturer_id, normalized_name FROM manufacturers WHERE name = ?",
            (candidate["canonical_name"],),
        ).fetchone()
        if legacy_row is None and canonical_row is not None:
            resolution_state = "merged"
        elif legacy_row is not None and canonical_row is not None:
            resolution_state = "pending"
        elif legacy_row is not None and canonical_row is None:
            resolution_state = "missing-canonical"
        else:
            resolution_state = "absent"
        normalization_rows.append(
            (
                candidate["legacy_name"],
                None if legacy_row is None else int(legacy_row[0]),
                None if legacy_row is None else str(legacy_row[1]),
                candidate["canonical_name"],
                None if canonical_row is None else int(canonical_row[0]),
                None if canonical_row is None else str(canonical_row[1]),
                resolution_state,
                candidate["reason"],
            )
        )
    outputs["manufacturer_normalization_candidates_csv"] = _export_rows(
        [
            "legacy_name",
            "legacy_manufacturer_id",
            "legacy_normalized_name",
            "canonical_name",
            "canonical_manufacturer_id",
            "canonical_normalized_name",
            "resolution_state",
            "reason",
        ],
        normalization_rows,
        export_dir / "manufacturer_normalization_candidates.csv",
    )
    return outputs


def parse_biglobe_armdata(html_text: str) -> list[ParsedTonearmRow]:
    parser = _TableRowParser()
    parser.feed(html_text)
    parsed: list[ParsedTonearmRow] = []
    for row in parser.rows:
        if len(row) != 6:
            continue
        cells = [normalize_space(cell) for cell in row]
        if cells[0].upper() == "MAKER":
            continue
        manufacturer, model, effective_length, overhang, offset_angle, null_points = cells
        if not manufacturer or not model:
            continue
        if manufacturer.startswith("**NOTES"):
            break
        if (normalize_name(manufacturer), normalize_name(model)) in BIGLOBE_EXCLUDED_MODELS:
            continue
        note_parts = []
        if any(ch in effective_length for ch in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"):
            note_parts.append(f"effective_length_raw={effective_length}")
        if any(ch in overhang for ch in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"):
            note_parts.append(f"overhang_raw={overhang}")
        if any(ch in offset_angle for ch in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"):
            note_parts.append(f"offset_angle_raw={offset_angle}")
        if "?" in effective_length or "?" in overhang or "?" in offset_angle or "?" in null_points:
            note_parts.append("contains uncertainty markers")
        parsed.append(
            ParsedTonearmRow(
                manufacturer=manufacturer,
                model=model,
                effective_length_mm=parse_first_float(effective_length),
                overhang_mm=parse_first_float(overhang),
                offset_angle_deg=parse_first_float(offset_angle),
                null_points=normalize_null_points_text(null_points),
                null_point_inner_mm=parse_null_point(null_points, 0),
                null_point_outer_mm=parse_null_point(null_points, 1),
                notes="; ".join(note_parts) if note_parts else None,
                raw_effective_length=effective_length,
                raw_overhang=overhang,
                raw_offset_angle=offset_angle,
                raw_null_points=null_points,
            )
        )
    return parsed


class _TableRowParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_tr = False
        self.in_cell = False
        self.current_row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self.in_tr = True
            self.current_row = []
        elif tag in {"td", "th"} and self.in_tr:
            self.in_cell = True
            self.current_row.append("")
        elif tag == "br" and self.in_cell and self.current_row:
            self.current_row[-1] += "\n"

    def handle_endtag(self, tag: str) -> None:
        if tag == "tr" and self.in_tr:
            self.rows.append(self.current_row)
            self.current_row = []
            self.in_tr = False
        elif tag in {"td", "th"}:
            self.in_cell = False

    def handle_data(self, data: str) -> None:
        if self.in_cell and self.current_row:
            self.current_row[-1] += data


def parse_first_float(text: str) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def parse_null_point(text: str, index: int) -> float | None:
    values = re.findall(r"\d+(?:\.\d+)?", text)
    if len(values) < index + 1:
        return None
    try:
        return float(values[index])
    except ValueError:
        return None


def normalize_null_points_text(text: str) -> str | None:
    cleaned = normalize_space(text)
    return cleaned or None


def normalize_null_alignment_type(text: str | None) -> str | None:
    cleaned = normalize_space(text or "")
    if not cleaned:
        return None
    if cleaned in NULL_ALIGNMENT_TYPE_VALUES:
        return cleaned
    lowered = cleaned.lower()
    has_baerwald = any(token in lowered for token in ("baerwald", "baewald", "lofgren a", "loefgren a"))
    has_loefgren = "lofgren" in lowered or "loefgren" in lowered
    has_stevenson = "stevenson" in lowered
    if "underhang" in lowered or "alternative alignment" in lowered:
        return "Other"
    if "baerwald default" in lowered or "baewald default" in lowered or "default baerwald" in lowered or "default baewald" in lowered:
        return "Baewald"
    if "stevenson default" in lowered or "default stevenson" in lowered:
        return "Stevenson"
    if "lofgren default" in lowered or "loefgren default" in lowered or "default lofgren" in lowered or "default loefgren" in lowered:
        return "Loefgren"
    if has_baerwald and has_loefgren and ("selectable" in lowered or "optional" in lowered or ' or ' in lowered):
        return "Other"
    if has_baerwald and has_stevenson and ("optional" in lowered or ' or ' in lowered):
        return "Other"
    if has_stevenson:
        return "Stevenson"
    if has_baerwald:
        return "Baewald"
    if has_loefgren:
        return "Loefgren"
    if lowered == "unknown":
        return "Unknown"
    if lowered == "other":
        return "Other"
    if "custom" in lowered and any(token in lowered for token in ("alignment", "geometry", "null point", "null-point", "protractor")):
        return "Other"
    return None


def derive_null_point_pair(
    effective_length_mm: float | None,
    mounting_distance_mm: float | None,
    overhang_mm: float | None,
    offset_angle_deg: float | None,
) -> tuple[float, float] | None:
    if effective_length_mm is None or offset_angle_deg is None:
        return None
    if mounting_distance_mm is None:
        if overhang_mm is None:
            return None
        mounting_distance_mm = effective_length_mm - overhang_mm
    sine_term = 2 * effective_length_mm * math.sin(math.radians(offset_angle_deg))
    product_term = effective_length_mm * effective_length_mm - mounting_distance_mm * mounting_distance_mm
    discriminant = sine_term * sine_term - 4 * product_term
    if discriminant < -1e-6:
        return None
    discriminant = max(0.0, discriminant)
    square_root = math.sqrt(discriminant)
    return ((sine_term - square_root) / 2, (sine_term + square_root) / 2)


def classify_null_alignment_type_from_pair(
    inner_null_mm: float | None,
    outer_null_mm: float | None,
) -> tuple[str, float] | None:
    if inner_null_mm is None or outer_null_mm is None:
        return None
    best_alignment = None
    best_distance = None
    for alignment_type, (reference_inner, reference_outer) in NULL_ALIGNMENT_REFERENCE_POINTS.items():
        distance = math.hypot(inner_null_mm - reference_inner, outer_null_mm - reference_outer)
        if best_distance is None or distance < best_distance:
            best_alignment = alignment_type
            best_distance = distance
    if best_alignment is None or best_distance is None:
        return None
    if best_distance <= NULL_ALIGNMENT_DISTANCE_TOLERANCE_MM:
        return best_alignment, best_distance
    return "Other", best_distance


def normalize_space(text: str) -> str:
    return " ".join(text.replace("\xa0", " ").replace("\u3000", " ").split())


def normalize_name(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def _coerce_workbook_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _canonicalize_workbook_manufacturer(name: str) -> str:
    direct_key = normalize_name(name)
    if direct_key in WORKBOOK_MANUFACTURER_ALIASES:
        return WORKBOOK_MANUFACTURER_ALIASES[direct_key]
    ascii_name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return WORKBOOK_MANUFACTURER_ALIASES.get(normalize_name(ascii_name), name)


def _is_probable_non_tonearm_workbook_row(manufacturer_name: str, model_name: str) -> bool:
    normalized_manufacturer = normalize_name(manufacturer_name)
    normalized_model = normalize_name(model_name)
    if (normalized_manufacturer, normalized_model) in WORKBOOK_EXCLUDED_MODELS:
        return True
    return "turntable" in normalized_model or "arm-wand" in normalized_model


def _delete_workbook_target_if_present(
    conn: sqlite3.Connection,
    *,
    workbook_name: str,
    manufacturer_name: str,
    model_name: str,
) -> None:
    normalized_manufacturer = normalize_name(manufacturer_name)
    normalized_model = normalize_name(model_name)
    queue_rows = conn.execute(
        """
        SELECT mrq.manufacturer_queue_id
        FROM manufacturer_research_queue mrq
        JOIN tonearm_research_targets trt ON trt.manufacturer_queue_id = mrq.manufacturer_queue_id
        WHERE mrq.normalized_name = ?
          AND lower(replace(replace(trt.model_name, '"', ''), '”', '')) = lower(replace(replace(?, '"', ''), '”', ''))
          AND (
              trt.source_hint = ?
              OR trt.notes LIKE ?
          )
        """,
        (
            normalized_manufacturer,
            model_name,
            f"user workbook: {workbook_name}",
            f"Imported from workbook {workbook_name}%",
        ),
    ).fetchall()
    conn.execute(
        """
        DELETE FROM tonearm_research_targets
        WHERE target_id IN (
            SELECT trt.target_id
            FROM tonearm_research_targets trt
            JOIN manufacturer_research_queue mrq ON mrq.manufacturer_queue_id = trt.manufacturer_queue_id
            WHERE mrq.normalized_name = ?
              AND lower(replace(replace(trt.model_name, '"', ''), '”', '')) = lower(replace(replace(?, '"', ''), '”', ''))
              AND (
                  trt.source_hint = ?
                  OR trt.notes LIKE ?
              )
        )
        """,
        (
            normalized_manufacturer,
            model_name,
            f"user workbook: {workbook_name}",
            f"Imported from workbook {workbook_name}%",
        ),
    )
    for (manufacturer_queue_id,) in queue_rows:
        remaining = conn.execute(
            "SELECT 1 FROM tonearm_research_targets WHERE manufacturer_queue_id = ? LIMIT 1",
            (manufacturer_queue_id,),
        ).fetchone()
        if remaining is None:
            conn.execute(
                """
                DELETE FROM manufacturer_research_queue
                WHERE manufacturer_queue_id = ?
                  AND discovery_source = ?
                  AND canonical_manufacturer_id IS NULL
                """,
                (manufacturer_queue_id, f"user-workbook:{workbook_name}"),
            )


def _build_workbook_gap_notes(
    *,
    workbook_name: str,
    source_manufacturer: str,
    effective_length: str,
    effective_length_source: str,
    mounting_distance: str,
    mounting_distance_source: str,
    workbook_notes: str,
) -> str:
    parts = [f"Imported from workbook {workbook_name}"]
    if source_manufacturer:
        parts.append(f"source_manufacturer={source_manufacturer}")
    if effective_length:
        parts.append(f"effective_length_raw={effective_length}")
    if effective_length_source:
        parts.append(f"effective_length_source={effective_length_source}")
    if mounting_distance:
        parts.append(f"mounting_distance_raw={mounting_distance}")
    if mounting_distance_source:
        parts.append(f"mounting_distance_source={mounting_distance_source}")
    if workbook_notes:
        parts.append(f"workbook_notes={workbook_notes}")
    return "; ".join(parts)


def _lookup_manufacturer_id(conn: sqlite3.Connection, name: str | None) -> int | None:
    if not name:
        return None
    row = conn.execute(
        "SELECT manufacturer_id FROM manufacturers WHERE normalized_name = ?",
        (normalize_name(name),),
    ).fetchone()
    return None if row is None else int(row[0])


def _upsert_manufacturer_research_seed(conn: sqlite3.Connection, seed: dict[str, str], now: str) -> None:
    manufacturer_name = seed["manufacturer_name"]
    normalized_name = normalize_name(manufacturer_name)
    canonical_match_name = seed.get("canonical_match_name") or manufacturer_name
    canonical_manufacturer_id = _lookup_manufacturer_id(conn, canonical_match_name)
    existing = conn.execute(
        "SELECT manufacturer_queue_id, created_at FROM manufacturer_research_queue WHERE normalized_name = ?",
        (normalized_name,),
    ).fetchone()
    if existing is None:
        conn.execute(
            """
            INSERT INTO manufacturer_research_queue(
                manufacturer_name, normalized_name, canonical_match_name, canonical_manufacturer_id,
                manufacturer_type, coverage_focus, priority_tier, status, discovery_source,
                search_terms, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                manufacturer_name,
                normalized_name,
                canonical_match_name,
                canonical_manufacturer_id,
                seed["manufacturer_type"],
                seed["coverage_focus"],
                seed["priority_tier"],
                seed["status"],
                seed.get("discovery_source"),
                seed.get("search_terms"),
                seed.get("notes"),
                now,
                now,
            ),
        )
        return

    conn.execute(
        """
        UPDATE manufacturer_research_queue
        SET manufacturer_name = ?,
            canonical_match_name = ?,
            canonical_manufacturer_id = ?,
            manufacturer_type = ?,
            coverage_focus = ?,
            priority_tier = ?,
            status = ?,
            discovery_source = ?,
            search_terms = ?,
            notes = ?,
            updated_at = ?
        WHERE manufacturer_queue_id = ?
        """,
        (
            manufacturer_name,
            canonical_match_name,
            canonical_manufacturer_id,
            seed["manufacturer_type"],
            seed["coverage_focus"],
            seed["priority_tier"],
            seed["status"],
            seed.get("discovery_source"),
            seed.get("search_terms"),
            seed.get("notes"),
            now,
            int(existing[0]),
        ),
    )


def _upsert_tonearm_research_target_seed(conn: sqlite3.Connection, seed: dict[str, str], now: str) -> None:
    queue_row = conn.execute(
        "SELECT manufacturer_queue_id FROM manufacturer_research_queue WHERE normalized_name = ?",
        (normalize_name(seed["manufacturer_name"]),),
    ).fetchone()
    if queue_row is None:
        return
    manufacturer_queue_id = int(queue_row[0])
    normalized_model_name = normalize_name(seed["model_name"])
    existing = conn.execute(
        "SELECT target_id FROM tonearm_research_targets WHERE manufacturer_queue_id = ? AND normalized_model_name = ?",
        (manufacturer_queue_id, normalized_model_name),
    ).fetchone()
    if existing is None:
        conn.execute(
            """
            INSERT INTO tonearm_research_targets(
                manufacturer_queue_id, model_name, normalized_model_name, target_group, target_type,
                priority_tier, status, source_hint, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                manufacturer_queue_id,
                seed["model_name"],
                normalized_model_name,
                seed.get("target_group"),
                seed["target_type"],
                seed["priority_tier"],
                seed["status"],
                seed.get("source_hint"),
                seed.get("notes"),
                now,
                now,
            ),
        )
        return

    conn.execute(
        """
        UPDATE tonearm_research_targets
        SET model_name = ?,
            target_group = ?,
            target_type = ?,
            priority_tier = ?,
            status = ?,
            source_hint = ?,
            notes = ?,
            updated_at = ?
        WHERE target_id = ?
        """,
        (
            seed["model_name"],
            seed.get("target_group"),
            seed["target_type"],
            seed["priority_tier"],
            seed["status"],
            seed.get("source_hint"),
            seed.get("notes"),
            now,
            int(existing[0]),
        ),
    )


def _insert_ingest_run(conn: sqlite3.Connection, started_at: str, notes: str) -> int:
    cursor = conn.execute(
        "INSERT INTO ingest_runs(started_at, status, notes) VALUES (?, 'running', ?)",
        (started_at, notes),
    )
    return int(cursor.lastrowid)


def _complete_ingest_run(conn: sqlite3.Connection, ingest_run_id: int) -> None:
    conn.execute(
        "UPDATE ingest_runs SET status='completed', completed_at=? WHERE ingest_run_id=?",
        (datetime.now(timezone.utc).replace(microsecond=0).isoformat(), ingest_run_id),
    )


def _insert_source(conn: sqlite3.Connection, snapshot: SourceSnapshot) -> int:
    cursor = conn.execute(
        """
        INSERT INTO sources(source_name, source_url, source_type, trust_level, retrieved_at, local_snapshot_path, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            snapshot.source_name,
            snapshot.source_url,
            snapshot.source_type,
            snapshot.trust_level,
            snapshot.retrieved_at,
            snapshot.local_snapshot_path,
            snapshot.notes,
        ),
    )
    return int(cursor.lastrowid)


def _upsert_model(conn: sqlite3.Connection, manufacturer: str, model: str, notes: str | None) -> int:
    manufacturer_norm = normalize_name(manufacturer)
    manufacturer_id = conn.execute(
        "SELECT manufacturer_id FROM manufacturers WHERE normalized_name = ?",
        (manufacturer_norm,),
    ).fetchone()
    if manufacturer_id is None:
        cursor = conn.execute(
            "INSERT INTO manufacturers(name, normalized_name, created_at) VALUES (?, ?, ?)",
            (manufacturer, manufacturer_norm, datetime.now(timezone.utc).replace(microsecond=0).isoformat()),
        )
        manufacturer_pk = int(cursor.lastrowid)
    else:
        manufacturer_pk = int(manufacturer_id[0])

    model_norm = normalize_name(model)
    existing = conn.execute(
        "SELECT tonearm_model_id FROM tonearm_models WHERE manufacturer_id = ? AND normalized_model_name = ?",
        (manufacturer_pk, model_norm),
    ).fetchone()
    if existing is not None:
        return int(existing[0])
    cursor = conn.execute(
        """
        INSERT INTO tonearm_models(manufacturer_id, model_name, normalized_model_name, display_name, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            manufacturer_pk,
            model,
            model_norm,
            f"{manufacturer} {model}",
            notes,
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        ),
    )
    return int(cursor.lastrowid)


def _insert_spec(
    conn: sqlite3.Connection,
    model_id: int,
    source_id: int,
    ingest_run_id: int | None,
    field_name: str,
    value_num: float | None,
    unit: str | None,
    raw_value_text: str | None,
    *,
    value_text: str | None = None,
    status: str | None = None,
    confidence: float | None = None,
    notes: str | None = None,
) -> None:
    if value_num is None and not value_text and not raw_value_text:
        return
    status = status or ("parsed" if value_num is not None or value_text else "raw_only")
    conn.execute(
        """
        INSERT INTO tonearm_specs(
            tonearm_model_id, source_id, ingest_run_id, field_name, value_num, value_text, unit, raw_value_text, status, confidence, notes, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            model_id,
            source_id,
            ingest_run_id,
            field_name,
            value_num,
            value_text,
            unit,
            raw_value_text,
            status,
            0.6 if confidence is None else confidence,
            notes,
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        ),
    )


def infer_null_alignment_type(
    model_name: str,
    model_notes: str | None,
    fact_rows: list[tuple[int, str, str, str, float | None, str | None, str | None, str | None, str]],
) -> tuple[str | None, int | None, str | None, float | None, str | None]:
    text_candidates: list[tuple[int | None, str, str]] = [(None, model_name, "secondary")]
    if model_notes:
        text_candidates.append((None, model_notes, "secondary"))
    field_values: dict[str, list[tuple[int, str, str, float | None, str | None, str | None, str | None]]] = {}
    for source_id, trust_level, source_type, field_name, value_num, value_text, raw_value_text, notes, status in fact_rows:
        if field_name == "null_alignment_type":
            continue
        field_values.setdefault(field_name, []).append((source_id, trust_level, source_type, value_num, value_text, raw_value_text, notes))
        for text_value in (value_text, raw_value_text, notes):
            if text_value:
                text_candidates.append((source_id, text_value, trust_level))

    for source_id, text_value, trust_level in sorted(text_candidates, key=lambda item: 0 if item[2] == "official" else 1):
        normalized_alignment_type = normalize_null_alignment_type(text_value)
        if normalized_alignment_type and normalized_alignment_type != "Unknown":
            return (
                normalized_alignment_type,
                source_id,
                normalize_space(text_value),
                0.95 if trust_level == "official" else 0.9,
                "Backfilled from explicit alignment-language evidence in model metadata or source text.",
            )

    inner_rows = field_values.get("null_point_inner_mm", [])
    outer_rows = field_values.get("null_point_outer_mm", [])

    def _classify_null_point_pair(preferred_trust: str | None) -> tuple[str, int | None, str, float, str] | None:
        filtered_inner_rows = [row for row in inner_rows if preferred_trust is None or row[1] == preferred_trust]
        filtered_outer_rows = [row for row in outer_rows if preferred_trust is None or row[1] == preferred_trust]
        if not filtered_inner_rows or not filtered_outer_rows:
            return None
        inner_source_id, _inner_trust, _inner_type, inner_value, *_ = filtered_inner_rows[-1]
        outer_source_id, _outer_trust, _outer_type, outer_value, *_ = filtered_outer_rows[-1]
        classified_pair = classify_null_alignment_type_from_pair(inner_value, outer_value)
        if classified_pair is None:
            return None
        alignment_type, distance = classified_pair
        return (
            alignment_type,
            inner_source_id or outer_source_id,
            f"{inner_value:.3f} / {outer_value:.3f} mm",
            0.9 if preferred_trust == "official" and alignment_type != "Other" else 0.85 if alignment_type != "Other" else 0.75,
            f"Backfilled from null-point pair proximity to reference alignments; nearest-reference distance {distance:.3f} mm.",
        )

    def _latest_value(field_name: str, preferred_trust: str | None) -> tuple[int | None, float | None] | None:
        rows = field_values.get(field_name, [])
        if preferred_trust is not None:
            rows = [row for row in rows if row[1] == preferred_trust]
        if not rows:
            return None
        return rows[-1][0], rows[-1][3]

    def _classify_derived_pair(preferred_trust: str | None) -> tuple[str, int | None, str, float, str] | None:
        effective_length_row = _latest_value("effective_length_mm", preferred_trust)
        offset_angle_row = _latest_value("offset_angle_deg", preferred_trust)
        mounting_distance_row = _latest_value("mounting_distance_mm", preferred_trust)
        overhang_row = _latest_value("overhang_mm", preferred_trust)
        derived_pair = derive_null_point_pair(
            None if effective_length_row is None else effective_length_row[1],
            None if mounting_distance_row is None else mounting_distance_row[1],
            None if overhang_row is None else overhang_row[1],
            None if offset_angle_row is None else offset_angle_row[1],
        )
        if derived_pair is None:
            return None
        classified_pair = classify_null_alignment_type_from_pair(*derived_pair)
        if classified_pair is None:
            return None
        alignment_type, distance = classified_pair
        source_candidates = [
            row[0]
            for row in (effective_length_row, mounting_distance_row, overhang_row, offset_angle_row)
            if row is not None and row[0] is not None
        ]
        source_id = source_candidates[-1] if source_candidates else None
        return (
            alignment_type,
            source_id,
            f"{derived_pair[0]:.3f} / {derived_pair[1]:.3f} mm",
            0.85 if preferred_trust == "official" and alignment_type != "Other" else 0.75 if alignment_type != "Other" else 0.65,
            f"Backfilled from derived null points using effective length, mounting distance/overhang, and offset angle; nearest-reference distance {distance:.3f} mm.",
        )

    for classifier in (
        lambda: _classify_null_point_pair("official"),
        lambda: _classify_derived_pair("official"),
        lambda: _classify_null_point_pair(None),
        lambda: _classify_derived_pair(None),
    ):
        classified = classifier()
        if classified is not None:
            return classified

    return None, None, None, None, None


def _ensure_null_alignment_type_defaults(conn: sqlite3.Connection, ingest_run_id: int | None = None) -> None:
    model_rows = conn.execute(
        """
        SELECT tm.tonearm_model_id, tm.model_name, tm.notes
        FROM tonearm_models tm
        ORDER BY tm.tonearm_model_id
        """
    ).fetchall()
    for model_id, model_name, model_notes in model_rows:
        existing_rows = conn.execute(
            """
            SELECT tonearm_spec_id, source_id, value_text, raw_value_text, status
            FROM tonearm_specs
            WHERE tonearm_model_id = ? AND field_name = 'null_alignment_type'
            ORDER BY tonearm_spec_id DESC
            """,
            (int(model_id),),
        ).fetchall()
        if any((row[4] or "") != "derived" for row in existing_rows):
            continue

        fact_rows = conn.execute(
            """
            SELECT ts.source_id, s.trust_level, s.source_type, ts.field_name, ts.value_num, ts.value_text, ts.raw_value_text, ts.notes, ts.status
            FROM tonearm_specs ts
            JOIN sources s ON s.source_id = ts.source_id
            WHERE ts.tonearm_model_id = ?
            ORDER BY ts.tonearm_spec_id
            """,
            (int(model_id),),
        ).fetchall()
        inferred_type, inferred_source_id, inferred_raw_text, inferred_confidence, inferred_notes = infer_null_alignment_type(
            model_name,
            model_notes,
            fact_rows,
        )
        target_value = inferred_type or "Unknown"
        target_raw = inferred_raw_text or target_value
        target_confidence = inferred_confidence if inferred_confidence is not None else 0.4
        target_source_id = inferred_source_id
        if target_source_id is None:
            first_source_row = next((row for row in fact_rows if row[0] is not None), None)
            if first_source_row is None:
                continue
            target_source_id = int(first_source_row[0])
        target_notes = inferred_notes or "Defaulted because the current source set does not explicitly specify a null alignment family."

        if existing_rows:
            conn.execute(
                """
                UPDATE tonearm_specs
                SET source_id = ?,
                    ingest_run_id = ?,
                    value_num = NULL,
                    value_text = ?,
                    unit = NULL,
                    raw_value_text = ?,
                    status = 'derived',
                    confidence = ?,
                    notes = ?
                WHERE tonearm_spec_id = ?
                """,
                (
                    int(target_source_id),
                    ingest_run_id,
                    target_value,
                    target_raw,
                    target_confidence,
                    target_notes,
                    int(existing_rows[0][0]),
                ),
            )
        else:
            _insert_spec(
                conn,
                int(model_id),
                int(target_source_id),
                ingest_run_id,
                "null_alignment_type",
                None,
                None,
                target_raw,
                value_text=target_value,
                status="derived",
                confidence=target_confidence,
                notes=target_notes,
            )


def _export_query(conn: sqlite3.Connection, query: str, path: Path) -> Path:
    cursor = conn.execute(query)
    columns = [item[0] for item in cursor.description]
    rows = cursor.fetchall()
    return _export_rows(columns, rows, path)


def _export_rows(columns: Iterable[str], rows: Iterable[Iterable[object]], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(list(columns))
        writer.writerows(rows)
    return path


def _build_readme(rows: Iterable[ParsedTonearmRow], snapshot: SourceSnapshot) -> str:
    row_count = len(list(rows)) if not isinstance(rows, list) else len(rows)
    return "\n".join(
        [
            "# Tonearm Database",
            "",
            "Generated local tonearm database artifacts.",
            "",
            "## Current hydration",
            f"- records hydrated: {row_count}",
            f"- primary source: {snapshot.source_name}",
            f"- source url: {snapshot.source_url}",
            f"- source type: {snapshot.source_type}",
            f"- trust level: {snapshot.trust_level}",
            "",
            "## Files",
            "- `tonearms.db` SQLite working database",
            "- `schema.sql` SQLite schema and preferred view",
            "- `sources/biglobe_armdata.html` local source snapshot",
            "- `exports/tonearms.csv` flattened preferred export",
            "- `exports/tonearm_specs.csv` field-level fact export",
            "- `exports/sources.csv` source registry export",
            "- `exports/manufacturers.csv` manufacturer export",
            "- `exports/tonearm_models.csv` model export",
            "",
            "## Notes",
            "- Current hydration is geometry-heavy because the source mostly provides effective length, overhang, offset angle, and null points.",
            "- Effective mass, cartridge range, and arm mount are left null until higher-detail sources are ingested.",
            "- Values with `?`, alternates, or commentary are preserved in raw text and parsed conservatively.",
        ]
    ) + "\n"
